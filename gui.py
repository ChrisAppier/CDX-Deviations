"""
gui.py
------
Tkinter desktop UI for WebFIRE Deviation Scanner.
Zero extra dependencies beyond requests + beautifulsoup4.

Usage:
    python3 gui.py
"""

import collections
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import zipfile
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import webfire_core as core

_CONFIG_PATH    = Path(__file__).parent / "config.json"
_DEFAULT_DL_DIR = Path(__file__).parent / "downloads"

VERSION = "1.0"

# ── Colours ────────────────────────────────────────────────────────────────
C_HEADER    = "#1a3a5c"
C_BTN_BLUE  = "#2563eb"
C_BTN_GRN   = "#16a34a"
C_BTN_AMB   = "#d97706"
C_DEV_BG    = "#ffe0e0"
C_DEV_FG    = "#990000"
C_ERR_BG    = "#fffbe6"
C_CACHED_BG = "#f0fff4"

_SORT_ASC  = " ▲"
_SORT_DESC = " ▼"


# ══════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("WebFIRE Deviation Scanner")
        self.geometry("1100x820")
        self.minsize(900, 600)
        self.configure(bg="#f0f2f5")

        self._session   = None
        self._reports   = []   # list[dict] from last search
        self._selected  = {}   # id -> bool (checkbox state)
        self._scan_rows = []
        self._scan_zip_dir = None   # set at scan time; Path to dir containing scanned ZIPs

        # Sort state for scan treeview (column name, ascending)
        self._sort_col = "result"
        self._sort_asc = True

        # Sort state for results treeview
        self._res_sort_col = None
        self._res_sort_asc = True

        # Download state
        self._dl_cancel = False

        # Filter vars (initialised properly in _build_scan_section)
        self._filter_var        = None
        self._filter_result_var = None

        # Section outer frames (set during _build_ui; used for ordering)
        self._search_card   = None
        self._results_card  = None
        self._dl_card       = None
        self._scan_card     = None

        self._active_folder: Path = self._load_active_folder()
        self._active_folder.mkdir(parents=True, exist_ok=True)
        self._folder_var: tk.StringVar | None = None  # assigned in _build_folder_bar

        self._build_ui()
        self._refresh_scan_button_state()  # H16: correct initial state

    # ── Working Folder persistence ─────────────────────────────────────────

    def _load_active_folder(self) -> Path:
        try:
            data = json.loads(_CONFIG_PATH.read_text(encoding="utf-8"))
            p = Path(data.get("active_folder", ""))
            if p and p.is_absolute():
                return p
        except Exception:
            pass
        return _DEFAULT_DL_DIR

    def _save_active_folder(self):
        try:
            _CONFIG_PATH.write_text(
                json.dumps({"active_folder": str(self._active_folder)}, indent=2),
                encoding="utf-8")
        except Exception:
            pass  # graceful degradation — folder still works this session

    def _set_active_folder(self, path: Path):
        """Single choke-point for all folder changes."""
        self._active_folder = path
        self._active_folder.mkdir(parents=True, exist_ok=True)
        self._save_active_folder()
        if self._folder_var:
            self._folder_var.set(str(self._active_folder))
        self._refresh_scan_button_state()

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Force light theme so dark-mode macOS doesn't swallow field text
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TEntry",
                         fieldbackground="white", foreground="#111111",
                         insertcolor="#111111", bordercolor="#cbd5e1",
                         lightcolor="#cbd5e1", darkcolor="#cbd5e1")
        style.configure("TCombobox",
                         fieldbackground="white", foreground="#111111",
                         selectbackground="white", selectforeground="#111111",
                         bordercolor="#cbd5e1")
        style.map("TCombobox",
                  fieldbackground=[("readonly", "white")],
                  foreground=[("readonly", "#111111")])
        style.configure("Treeview",
                         background="white", foreground="#111111",
                         fieldbackground="white", rowheight=22)
        style.configure("Treeview.Heading",
                         background="#eef1f5", foreground="#1e293b",
                         font=("Helvetica", 9, "bold"), relief="flat")
        style.map("Treeview",
                  background=[("selected", "#2563eb")],
                  foreground=[("selected", "white")])
        style.configure("TScrollbar",
                         troughcolor="#e2e8f0", background="#94a3b8")
        style.configure("Horizontal.TProgressbar",
                         troughcolor="#e2e8f0", background="#16a34a")
        style.configure("TProgressbar",
                         troughcolor="#e2e8f0", background="#16a34a")

        # ── Button styles — tk.Button ignores bg/fg on macOS Aqua; ttk+clam works
        style.configure("Blue.TButton",
                        background=C_BTN_BLUE, foreground="white",
                        font=("Helvetica", 10, "bold"), relief="flat",
                        borderwidth=0, padding=(6, 5))
        style.map("Blue.TButton",
                  background=[("active", "#1d4ed8"), ("disabled", "#93c5fd")],
                  foreground=[("active", "white"), ("disabled", "#dbeafe")])
        style.configure("Green.TButton",
                        background=C_BTN_GRN, foreground="white",
                        font=("Helvetica", 9, "bold"), relief="flat",
                        borderwidth=0, padding=(6, 3))
        style.map("Green.TButton",
                  background=[("active", "#15803d"), ("disabled", "#86efac")],
                  foreground=[("active", "white"), ("disabled", "#dcfce7")])
        style.configure("Amber.TButton",
                        background=C_BTN_AMB, foreground="white",
                        font=("Helvetica", 10, "bold"), relief="flat",
                        borderwidth=0, padding=(6, 4))
        style.map("Amber.TButton",
                  background=[("active", "#b45309")],
                  foreground=[("active", "white")])
        style.configure("DarkGrey.TButton",
                        background="#4b5563", foreground="white",
                        font=("Helvetica", 9, "bold"), relief="flat",
                        borderwidth=0, padding=(6, 4))
        style.map("DarkGrey.TButton",
                  background=[("active", "#374151"), ("disabled", "#9ca3af")],
                  foreground=[("active", "white"), ("disabled", "#f3f4f6")])
        style.configure("Red.TButton",
                        background="#dc2626", foreground="white",
                        font=("Helvetica", 9, "bold"), relief="flat",
                        borderwidth=0, padding=(6, 2))
        style.map("Red.TButton",
                  background=[("active", "#b91c1c")],
                  foreground=[("active", "white")])
        style.configure("Header.TButton",
                        background="#2d6a9f", foreground="white",
                        font=("Helvetica", 11, "bold"), relief="flat",
                        borderwidth=0, padding=(8, 2))
        style.map("Header.TButton",
                  background=[("active", "#1d4ed8")],
                  foreground=[("active", "white")])
        style.configure("Grey.TButton",
                        background="#e5e7eb", foreground="#374151",
                        font=("Helvetica", 9), relief="flat",
                        borderwidth=0, padding=(6, 4))
        style.map("Grey.TButton",
                  background=[("active", "#d1d5db"), ("disabled", "#e5e7eb")],
                  foreground=[("active", "#374151"), ("disabled", "#6b7280")])

        # ── Title bar
        hdr = tk.Frame(self, bg=C_HEADER, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="WebFIRE Deviation Scanner",
                 bg=C_HEADER, fg="white",
                 font=("Helvetica", 15, "bold")).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text="EPA CDX / WebFIRE Report Search & Analysis",
                 bg=C_HEADER, fg="#93b8d8",
                 font=("Helvetica", 10)).pack(side="left", padx=4, pady=10)
        # M27 — Help/About button
        ttk.Button(hdr, text=" ? ", style="Header.TButton",
                   cursor="hand2",
                   command=self._show_help).pack(side="right", padx=12, pady=10)

        # ── Working Folder bar
        self._build_folder_bar()

        # ── Scrollable main canvas
        outer = tk.Frame(self, bg="#f0f2f5")
        outer.pack(fill="both", expand=True)
        self._canvas = tk.Canvas(outer, bg="#f0f2f5", highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._main = tk.Frame(self._canvas, bg="#f0f2f5", padx=16, pady=12)
        self._canvas_win = self._canvas.create_window(
            (0, 0), window=self._main, anchor="nw")

        self._main.bind("<Configure>", self._on_frame_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)
        # L22 — mousewheel only fires on canvas when cursor is not over a treeview/text
        self._canvas.bind("<MouseWheel>", self._on_mousewheel)

        # ── Steps
        self._build_search_section()
        self._build_results_section()
        self._build_download_section()
        self._build_scan_section()

        # ── Status bar
        self._status_var = tk.StringVar(value="Ready")
        sb = tk.Frame(self, bg="#e2e8f0", height=26)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        tk.Label(sb, textvariable=self._status_var,
                 bg="#e2e8f0", fg="#475569",
                 font=("Helvetica", 10), anchor="w").pack(
                     side="left", padx=10, pady=3)

        # H25 — Global keyboard shortcuts
        self.bind_all("<Return>",      self._kb_search)
        self.bind_all("<Command-a>",   self._kb_select_all)
        self.bind_all("<Command-s>",   self._kb_export)
        self.bind_all("<Escape>",      self._kb_cancel)

    def _build_folder_bar(self):
        self._folder_var = tk.StringVar(value=str(self._active_folder))

        bar = tk.Frame(self, bg="#e8edf3", pady=5, padx=14)
        bar.pack(fill="x")  # packs between header and scrollable canvas

        ttk.Button(bar, text="  \U0001f4c2  Change Folder\u2026  ",
                   style="Blue.TButton", cursor="hand2",
                   command=self._pick_working_folder).pack(side="left")

        tk.Label(bar, text="Working Folder:",
                 bg="#e8edf3", fg="#374151",
                 font=("Helvetica", 9, "bold")).pack(side="left", padx=(10, 0))

        tk.Label(bar, textvariable=self._folder_var,
                 bg="#e8edf3", fg="#1e3a5c",
                 font=("Helvetica", 9), anchor="w").pack(
                     side="left", padx=(6, 0), fill="x", expand=True)

    def _pick_working_folder(self):
        chosen = filedialog.askdirectory(
            title="Select Working Folder \u2014 downloads will go here; scans will read from here",
            initialdir=str(self._active_folder))
        if chosen:
            self._set_active_folder(Path(chosen))

    def _card(self, parent, step, title, hidden=False, collapsed=False):
        outer = tk.Frame(parent, bg="#f0f2f5")
        outer.pack(fill="x", pady=(0, 12))

        # ── Clickable header ──
        hdr = tk.Frame(outer, bg=C_HEADER, cursor="hand2")
        hdr.pack(fill="x")
        badge = tk.Label(hdr, text=str(step), bg="#2d6a9f", fg="white",
                         font=("Helvetica", 10, "bold"),
                         width=2, relief="flat", cursor="hand2")
        badge.pack(side="left", padx=(8, 6), pady=6)
        title_lbl = tk.Label(hdr, text=title, bg=C_HEADER, fg="white",
                             font=("Helvetica", 10, "bold"), cursor="hand2")
        title_lbl.pack(side="left", pady=6)
        arrow_lbl = tk.Label(hdr, text=" ▼ ", bg=C_HEADER, fg="white",
                              font=("Helvetica", 10), cursor="hand2")
        arrow_lbl.pack(side="right", padx=6, pady=6)

        # ── Body ──
        body = tk.Frame(outer, bg="white", relief="flat", bd=0,
                        highlightbackground="#dde1e7", highlightthickness=1)
        body.pack(fill="x")

        def _toggle(_event=None):
            if body.winfo_ismapped():
                body.pack_forget()
                arrow_lbl.configure(text=" ▶ ")
            else:
                body.pack(fill="x")
                arrow_lbl.configure(text=" ▼ ")

        def _expand():
            if not body.winfo_ismapped():
                body.pack(fill="x")
                arrow_lbl.configure(text=" ▼ ")

        for w in (hdr, badge, title_lbl, arrow_lbl):
            w.bind("<Button-1>", _toggle)

        if hidden:
            outer.pack_forget()
        if collapsed:
            body.pack_forget()
            arrow_lbl.configure(text=" ▶ ")

        return outer, body, _expand

    # ── Step 1: Search ─────────────────────────────────────────────────────

    def _build_search_section(self):
        self._search_card, body, _ = self._card(self._main, 1, "SEARCH PARAMETERS")
        body.configure(padx=14, pady=12)

        # Row 0
        r0 = tk.Frame(body, bg="white")
        r0.pack(fill="x", pady=(0, 8))

        self._fac_var = self._labeled_entry(r0, "Facility Name", 0, width=22)
        self._org_var = self._labeled_entry(r0, "Organization",  1, width=22)

        # State dropdown — H2: show full label, not just code
        tk.Label(r0, text="State", bg="white", fg="#111111",
                 font=("Helvetica", 9, "bold")).grid(row=0, column=4, sticky="w", padx=(12, 4))
        state_values = [f"{v}  —  {l}" for v, l in core.STATE_OPTIONS]
        self._state_var = tk.StringVar(value=state_values[0])
        state_cb = ttk.Combobox(r0, textvariable=self._state_var, width=22,
                                state="readonly")
        state_cb["values"] = state_values
        state_cb.current(0)
        state_cb.grid(row=1, column=4, padx=(12, 4), sticky="w")

        self._city_var = self._labeled_entry(r0, "City", 5, width=14)

        # Row 1
        r1 = tk.Frame(body, bg="white")
        r1.pack(fill="x", pady=(0, 8))

        self._zip_var   = self._labeled_entry(r1, "ZIP Code",               0, width=12)
        self._start_var = self._labeled_entry(r1, "Start Date (MM/DD/YYYY)", 1, width=16)
        self._end_var   = self._labeled_entry(r1, "End Date (MM/DD/YYYY)",   2, width=16)

        # CFR Part
        tk.Label(r1, text="CFR Part", bg="white", fg="#111111",
                 font=("Helvetica", 9, "bold")).grid(row=0, column=5, sticky="w", padx=(12, 4))
        self._cfrpart_var = tk.StringVar(value="All")
        cfr_cb = ttk.Combobox(r1, textvariable=self._cfrpart_var, width=22, state="readonly")
        cfr_cb["values"] = [l for _, l in core.CFR_PART_OPTIONS]
        cfr_cb.current(0)
        cfr_cb.grid(row=1, column=5, padx=(12, 4), sticky="w")

        # L5 — CFR Subpart with format hint
        tk.Label(r1, text="CFR Subpart  (e.g. NNNN)", bg="white", fg="#111111",
                 font=("Helvetica", 9, "bold")).grid(row=0, column=6, sticky="w", padx=(12, 4))
        self._cfrsub_var = tk.StringVar()
        sub_entry = tk.Entry(r1, textvariable=self._cfrsub_var, width=12,
                             bg="white", fg="#111111", insertbackground="#111111",
                             relief="solid", bd=1,
                             highlightthickness=1, highlightcolor="#2563eb",
                             highlightbackground="#cbd5e1",
                             font=("Helvetica", 10))
        sub_entry.grid(row=1, column=6, padx=(12, 4), sticky="w")

        # Button row — M3: add Clear button, M4: add search spinner
        btn_row = tk.Frame(body, bg="white")
        btn_row.pack(fill="x", pady=(4, 0))
        self._btn_search = ttk.Button(
            btn_row, text="  Search WebFIRE  ",
            style="Blue.TButton", cursor="hand2",
            command=self._do_search)
        self._btn_search.pack(side="left")
        ttk.Button(btn_row, text="Clear", style="Grey.TButton",
                   cursor="hand2",
                   command=self._reset_search).pack(side="left", padx=8)
        # M4 — indeterminate spinner shown during search
        self._search_spinner = ttk.Progressbar(btn_row, mode="indeterminate",
                                               length=120)
        self._search_lbl = tk.Label(btn_row, text="", bg="white",
                                    fg="#374151", font=("Helvetica", 9))
        self._search_lbl.pack(side="left", padx=12)

    def _labeled_entry(self, parent, label, col, width=18):
        tk.Label(parent, text=label, bg="white", fg="#111111",
                 font=("Helvetica", 9, "bold")).grid(
                     row=0, column=col*2, sticky="w",
                     padx=(0 if col == 0 else 12, 4))
        var = tk.StringVar()
        e = tk.Entry(parent, textvariable=var, width=width,
                     bg="white", fg="#111111", insertbackground="#111111",
                     relief="solid", bd=1,
                     highlightthickness=1, highlightcolor="#2563eb",
                     highlightbackground="#cbd5e1",
                     font=("Helvetica", 10))
        e.grid(row=1, column=col*2, padx=(0 if col == 0 else 12, 4), sticky="w")
        return var

    # ── Step 2: Results ────────────────────────────────────────────────────

    def _build_results_section(self):
        self._results_card, body, self._expand_results = self._card(
            self._main, 2, "SEARCH RESULTS", collapsed=True)
        body.configure(padx=0, pady=0)

        # Toolbar
        tb = tk.Frame(body, bg="#f8fafc", pady=6, padx=10)
        tb.pack(fill="x")
        self._res_count_lbl = tk.Label(tb, text="", bg="#f8fafc",
                                       fg="#111111", font=("Helvetica", 9))
        self._res_count_lbl.pack(side="left")
        # M8 — cached row legend
        tk.Label(tb, text="  \u2588 = previously downloaded",
                 bg="#f8fafc", fg="#16a34a",
                 font=("Helvetica", 8)).pack(side="left", padx=(8, 0))
        ttk.Button(tb, text="Select All", style="Grey.TButton",
                   cursor="hand2",
                   command=self._select_all).pack(side="left", padx=(12, 4))
        ttk.Button(tb, text="Clear", style="Grey.TButton",
                   cursor="hand2",
                   command=self._select_none).pack(side="left", padx=4)
        self._btn_download = ttk.Button(
            tb, text="  \u2193  Download Selected  ",
            style="Green.TButton", cursor="hand2",
            command=self._do_download)
        self._btn_download.pack(side="right", padx=6)

        # Treeview
        cols = ("chk", "facility", "city", "st",
                "date", "type", "subtype", "pollutants", "status")
        self._res_tree = ttk.Treeview(
            body, columns=cols, show="headings", height=10, selectmode="none")

        widths = {"chk": 28, "facility": 200, "city": 100,
                  "st": 36, "date": 100, "type": 52, "subtype": 70,
                  "pollutants": 160, "status": 90}
        # H7 — sortable headings for results tree
        self._res_heads = {"chk": "", "facility": "Facility",
                           "city": "City", "st": "St", "date": "Date",
                           "type": "Type", "subtype": "Sub",
                           "pollutants": "Pollutants", "status": "Status"}
        for c in cols:
            self._res_tree.heading(c, text=self._res_heads[c],
                                   command=lambda col=c: self._on_res_sort(col))
            self._res_tree.column(c, width=widths[c],
                                  stretch=(c in ("facility", "pollutants")))

        self._res_tree.tag_configure("cached", background=C_CACHED_BG)

        vsb2 = ttk.Scrollbar(body, orient="vertical",
                              command=self._res_tree.yview)
        self._res_tree.configure(yscrollcommand=vsb2.set)
        self._res_tree.pack(side="left", fill="both", expand=True)
        vsb2.pack(side="right", fill="y")

        # H6 — full row click toggles checkbox
        self._res_tree.bind("<ButtonRelease-1>", self._toggle_checkbox)

        # M9 — empty results placeholder
        self._res_empty_lbl = tk.Label(
            body, text="No reports found — try broadening your search criteria.",
            bg="white", fg="#9ca3af", font=("Helvetica", 10), pady=14)

        # L22 — suppress canvas mousewheel when cursor is over the results tree
        self._bind_scroll_suppression(self._res_tree)

    # ── Step 3: Download progress ──────────────────────────────────────────

    def _build_download_section(self):
        self._dl_card, body, self._expand_dl = self._card(
            self._main, 3, "DOWNLOAD PROGRESS", collapsed=True)
        body.configure(padx=14, pady=10)

        prog_row = tk.Frame(body, bg="white")
        prog_row.pack(fill="x", pady=(0, 6))
        self._dl_bar = ttk.Progressbar(prog_row, mode="determinate", length=500)
        self._dl_bar.pack(side="left", fill="x", expand=True)
        self._dl_lbl = tk.Label(prog_row, text="", bg="white",
                                fg="#111111", font=("Helvetica", 9), width=10)
        self._dl_lbl.pack(side="left", padx=8)
        self._btn_cancel_dl = ttk.Button(
            prog_row, text="Cancel",
            style="Red.TButton", cursor="hand2",
            command=self._cancel_download)
        # Hidden until a download is in progress
        self._dl_cancel = False

        # M13 — current file label
        self._dl_file_lbl = tk.Label(body, text="", bg="white", fg="#374151",
                                     font=("Helvetica", 9), anchor="w")
        self._dl_file_lbl.pack(fill="x", pady=(0, 4))

        # H11 — log with both scrollbars; L14 — height 8
        log_frame = tk.Frame(body, bg="white")
        log_frame.pack(fill="x")
        self._dl_log = tk.Text(log_frame, height=8, state="disabled",
                               bg="#f8fafc", fg="#111111",
                               font=("Courier", 9), relief="flat",
                               wrap="none", borderwidth=0,
                               insertbackground="#111111")
        dl_log_vsb = ttk.Scrollbar(log_frame, orient="vertical",
                                   command=self._dl_log.yview)
        dl_log_hsb = ttk.Scrollbar(log_frame, orient="horizontal",
                                   command=self._dl_log.xview)
        self._dl_log.configure(yscrollcommand=dl_log_vsb.set,
                               xscrollcommand=dl_log_hsb.set)
        dl_log_vsb.pack(side="right", fill="y")
        dl_log_hsb.pack(side="bottom", fill="x")
        self._dl_log.pack(side="left", fill="both", expand=True)

        # L22 — suppress canvas scroll when over log
        self._bind_scroll_suppression(self._dl_log)

    # ── Step 4: Scan ───────────────────────────────────────────────────────

    def _build_scan_section(self):
        self._scan_card, body, self._expand_scan = self._card(
            self._main, 4, "DEVIATION SCAN", collapsed=True)
        body.configure(padx=0, pady=0)

        # ── Toolbar ─────────────────────────────────────────────────────────
        tb = tk.Frame(body, bg="#f8fafc", pady=6, padx=10)
        tb.pack(fill="x")
        # H16 — scan button starts as no-op (cursor=arrow); enabled when ZIPs exist
        self._btn_scan = ttk.Button(
            tb, text="  \U0001f50d  Scan for Deviations  ",
            style="Amber.TButton", cursor="arrow",
            command=lambda: None)
        self._btn_scan.pack(side="left")
        self._btn_export = ttk.Button(
            tb, text="  Export CSV  ",
            style="Grey.TButton", cursor="hand2",
            command=self._do_export, state="disabled")
        self._btn_export.pack(side="left", padx=4)
        self._btn_export_xlsx = ttk.Button(
            tb, text="  Export XLSX  ",
            style="Grey.TButton", cursor="hand2",
            command=self._do_export_xlsx, state="disabled")
        self._btn_export_xlsx.pack(side="left", padx=4)
        self._btn_extract = ttk.Button(
            tb, text="  Extract Files\u2026  ",
            style="Grey.TButton", cursor="hand2",
            command=self._do_extract_files, state="disabled")
        self._btn_extract.pack(side="left", padx=4)

        # ── Filter bar ───────────────────────────────────────────────────────
        fb = tk.Frame(body, bg="#eef1f5", pady=5, padx=10)
        fb.pack(fill="x")
        tk.Label(fb, text="Search (facility, citation, sheet, notes — shows matching reports):",
                 bg="#eef1f5", fg="#111111",
                 font=("Helvetica", 9, "bold")).pack(side="left")
        self._filter_var = tk.StringVar()
        self._filter_var.trace_add("write", lambda *_: self._apply_scan_filter())
        tk.Entry(fb, textvariable=self._filter_var, width=28,
                 bg="white", fg="#111111", insertbackground="#111111",
                 relief="solid", bd=1, font=("Helvetica", 10)
                 ).pack(side="left", padx=(4, 12))
        tk.Label(fb, text="Show:", bg="#eef1f5", fg="#111111",
                 font=("Helvetica", 9, "bold")).pack(side="left")
        self._filter_result_var = tk.StringVar(value="All")
        result_cb = ttk.Combobox(fb, textvariable=self._filter_result_var, width=14,
                                  state="readonly",
                                  # M17 — "Review PDF" → "Manual Review"
                                  values=["All", "Deviations", "Manual Review", "Pass", "Errors"])
        result_cb.pack(side="left", padx=4)
        result_cb.bind("<<ComboboxSelected>>", lambda _: self._apply_scan_filter())
        ttk.Button(fb, text="Clear", style="Grey.TButton",
                   cursor="hand2",
                   command=self._clear_filter).pack(side="left", padx=8)

        # L24 — fallback warning as full-width strip below filter bar;
        # always packed — collapses to 0 height when label text is empty
        self._fallback_strip = tk.Frame(body, bg="#fef3c7")
        self._fallback_strip.pack(fill="x")
        self._fallback_lbl   = tk.Label(self._fallback_strip,
                                        text="", bg="#fef3c7",
                                        fg="#92400e", font=("Helvetica", 9),
                                        anchor="w", padx=10)
        self._fallback_lbl.pack(fill="x")

        # ── Scan progress bar (hidden until scan runs) ───────────────────────
        self._scan_prog_frame = tk.Frame(body, bg="white", padx=14, pady=6)
        self._scan_prog_frame.pack(fill="x")
        self._scan_prog_frame.pack_forget()
        sp_row = tk.Frame(self._scan_prog_frame, bg="white")
        sp_row.pack(fill="x")
        self._scan_bar = ttk.Progressbar(sp_row, mode="determinate", length=500)
        self._scan_bar.pack(side="left", fill="x", expand=True)
        self._scan_lbl = tk.Label(sp_row, text="", bg="white",
                                  fg="#111111", font=("Helvetica", 9), width=12)
        self._scan_lbl.pack(side="left", padx=8)

        # ── Summary label — M20: larger, bold, more prominent ────────────────
        self._scan_summary = tk.Label(body, text="", bg="white",
                                      font=("Helvetica", 12, "bold"),
                                      anchor="w", padx=14, pady=8)
        self._scan_summary.pack(fill="x")
        self._scan_summary.pack_forget()

        # ── Results treeview ─────────────────────────────────────────────────
        self._tree_frame = tk.Frame(body, bg="white")
        self._tree_frame.pack(fill="both", expand=True)

        scols = ("frs_id", "result", "facility", "st", "date", "type",
                 "location", "pollutant", "measured", "limit",
                 "pct", "notes")
        self._scan_tree = ttk.Treeview(
            self._tree_frame, columns=scols, show="tree headings", height=14)

        self._scan_tree.heading("#0", text="")
        self._scan_tree.column("#0", width=20, stretch=False, minwidth=20)

        # H15 — headings with sort direction tracking
        self._scan_heads = {
            "frs_id":   "FRS ID",   "result":   "Result",
            "facility": "Facility", "st":       "St",
            "date":     "Date",     "type":     "Type",
            "location": "Location / Sheet",
            "pollutant":"Pollutant / Citation",
            "measured": "Measured", "limit":    "Limit",
            "pct":      "% of Lim", "notes":    "Notes",
        }
        swidths = {"frs_id": 80, "result": 80, "facility": 180, "st": 36,
                   "date": 90, "type": 52, "location": 90, "pollutant": 180,
                   "measured": 80, "limit": 70, "pct": 68,
                   "notes": 300}
        for c in scols:
            self._scan_tree.heading(
                c, text=self._scan_heads[c],
                command=lambda col=c: self._on_scan_sort(col))
            self._scan_tree.column(
                c, width=swidths[c],
                stretch=(c in ("facility", "pollutant", "notes")))

        self._scan_tree.tag_configure("deviation",
                                      background=C_DEV_BG, foreground=C_DEV_FG)
        self._scan_tree.tag_configure("error",       background=C_ERR_BG)
        self._scan_tree.tag_configure("nolimit",     foreground="#6b7280")
        self._scan_tree.tag_configure("manualreview", background="#fff7e6",
                                      foreground="#92400e")
        self._scan_tree.tag_configure("fallback",    foreground="#b45309")

        scan_hsb = ttk.Scrollbar(self._tree_frame, orient="horizontal",
                                  command=self._scan_tree.xview)
        scan_vsb = ttk.Scrollbar(self._tree_frame, orient="vertical",
                                  command=self._scan_tree.yview)
        self._scan_tree.configure(yscrollcommand=scan_vsb.set,
                                  xscrollcommand=scan_hsb.set)
        scan_hsb.pack(side="bottom", fill="x")
        self._scan_tree.pack(side="left", fill="both", expand=True)
        scan_vsb.pack(side="right", fill="y")

        # M18 — double-click opens the report document
        self._scan_tree.bind("<Double-1>", self._open_scan_result_folder)

        # L22/L28 — suppress canvas mousewheel when over scan treeview
        self._bind_scroll_suppression(self._scan_tree)

        # Update sort heading indicator for default sort
        self._update_scan_sort_heading()

    # ── Canvas / scroll helpers ────────────────────────────────────────────

    def _on_frame_configure(self, _e):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_configure(self, e):
        self._canvas.itemconfig(self._canvas_win, width=e.width)

    def _on_mousewheel(self, e):
        self._canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

    # ── Checkbox helpers ───────────────────────────────────────────────────

    def _toggle_checkbox(self, event):
        # H6 — toggle on any cell click, not just the checkbox column
        region = self._res_tree.identify_region(event.x, event.y)
        iid    = self._res_tree.identify_row(event.y)
        if region != "cell" or not iid:
            return
        self._selected[iid] = not self._selected.get(iid, False)
        self._refresh_checkbox(iid)

    def _refresh_checkbox(self, iid):
        vals = list(self._res_tree.item(iid, "values"))
        vals[0] = "\u2611" if self._selected.get(iid) else "\u2610"
        self._res_tree.item(iid, values=vals)

    def _select_all(self):
        for iid in self._res_tree.get_children():
            self._selected[iid] = True
            self._refresh_checkbox(iid)

    def _select_none(self):
        for iid in self._res_tree.get_children():
            self._selected[iid] = False
            self._refresh_checkbox(iid)

    # ── Status / card helpers ──────────────────────────────────────────────

    def _set_status(self, msg):
        self._status_var.set(msg)

    def _show_card(self, card_outer):
        """Show a card section, inserting it in the correct 1→2→3→4 order."""
        sections = [self._search_card, self._results_card,
                    self._dl_card, self._scan_card]
        idx = sections.index(card_outer)

        # Find the nearest preceding section that is currently pack-managed
        after_widget = None
        for i in range(idx - 1, -1, -1):
            try:
                sections[i].pack_info()
                after_widget = sections[i]
                break
            except tk.TclError:
                pass

        if after_widget is not None:
            card_outer.pack(after=after_widget, fill="x", pady=(0, 12))
        else:
            # Fall back to packing before the nearest following visible section
            before_widget = None
            for i in range(idx + 1, len(sections)):
                try:
                    sections[i].pack_info()
                    before_widget = sections[i]
                    break
                except tk.TclError:
                    pass
            if before_widget is not None:
                card_outer.pack(before=before_widget, fill="x", pady=(0, 12))
            else:
                card_outer.pack(fill="x", pady=(0, 12))

    # H16 — enable/disable scan button based on whether ZIPs exist
    def _refresh_scan_button_state(self):
        has_zips = any(self._active_folder.glob("*.zip"))
        if has_zips:
            self._btn_scan.configure(command=self._do_scan, cursor="hand2")
        else:
            self._btn_scan.configure(command=lambda: None, cursor="arrow")

    # ── Sort helpers ───────────────────────────────────────────────────────

    def _update_scan_sort_heading(self):
        """Refresh scan treeview headings to show ▲/▼ on active column."""
        indicator = _SORT_ASC if self._sort_asc else _SORT_DESC
        for c in self._scan_heads:
            base = self._scan_heads[c]
            text = base + indicator if c == self._sort_col else base
            self._scan_tree.heading(c, text=text)

    def _update_res_sort_heading(self):
        """Refresh results treeview headings to show ▲/▼ on active column."""
        if self._res_sort_col is None:
            return
        indicator = _SORT_ASC if self._res_sort_asc else _SORT_DESC
        for c in self._res_heads:
            base = self._res_heads[c]
            text = base + indicator if c == self._res_sort_col else base
            self._res_tree.heading(c, text=text,
                                   command=lambda col=c: self._on_res_sort(col))

    # ─────────────────────────────────────────────────────────────────────
    # Actions
    # ─────────────────────────────────────────────────────────────────────

    # ── Keyboard shortcut handlers (H25) ───────────────────────────────────

    def _kb_search(self, event=None):
        if self._btn_search["state"] != "disabled":
            self._do_search()

    def _kb_select_all(self, _event=None):
        if self._res_tree.winfo_ismapped() and self._reports:
            self._select_all()
            return "break"

    def _kb_export(self, _event=None):
        if self._scan_rows:
            self._do_export()
            return "break"

    def _kb_cancel(self, _event=None):
        if not self._dl_cancel and self._btn_cancel_dl.winfo_ismapped():
            self._cancel_download()

    # ── Search ─────────────────────────────────────────────────────────────

    def _reset_search(self):
        """M3 — clear all search fields."""
        self._fac_var.set("")
        self._org_var.set("")
        self._city_var.set("")
        self._zip_var.set("")
        self._start_var.set("")
        self._end_var.set("")
        self._cfrsub_var.set("")
        state_values = [f"{v}  —  {l}" for v, l in core.STATE_OPTIONS]
        self._state_var.set(state_values[0])
        self._cfrpart_var.set("All")
        self._search_lbl.configure(text="")

    def _do_search(self, _event=None):
        # H1 — validate date fields before firing
        date_pat = re.compile(r"^\d{2}/\d{2}/\d{4}$")
        for label, var in (("Start Date", self._start_var), ("End Date", self._end_var)):
            val = var.get().strip()
            if val and not date_pat.match(val):
                messagebox.showerror("Invalid Date",
                                     f"{label} must be in MM/DD/YYYY format.\nGot: {val!r}")
                return

        self._btn_search.configure(state="disabled", text="  Searching…  ")
        self._search_lbl.configure(text="")
        # M4 — show spinner
        self._search_spinner.pack(side="left", padx=(0, 8))
        self._search_spinner.start(12)
        self._set_status("Connecting to WebFIRE…")

        def worker():
            try:
                if self._session is None:
                    self._session = core.build_session()
                params = {
                    "facility":     self._fac_var.get().strip(),
                    "organization": self._org_var.get().strip(),
                    "state":        self._state_var.get().split("  —  ")[0].strip(),
                    "city":         self._city_var.get().strip(),
                    "zip":          self._zip_var.get().strip(),
                    "startdate":    self._start_var.get().strip(),
                    "enddate":      self._end_var.get().strip(),
                    "CFRpart":      _cfr_value(self._cfrpart_var.get()),
                    "CFRSubpart":   self._cfrsub_var.get().strip(),
                }
                reports = core.search(self._session, params)
                self.after(0, lambda: self._on_search_done(reports))
            except Exception as exc:
                self.after(0, lambda: self._on_search_error(str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _on_search_done(self, reports):
        self._stop_search_spinner()
        self._expand_results()

        self._reports  = reports
        self._selected = {}

        # Populate tree
        self._res_tree.delete(*self._res_tree.get_children())
        for r in reports:
            cached = (self._active_folder / f"{r['id']}.zip").exists()
            r["downloaded"] = cached
            self._insert_result_row(r, checked=cached, cached=cached)
            self._selected[r["id"]] = cached  # pre-select cached rows

        n = len(reports)
        self._res_count_lbl.configure(
            text=f"{n} report{'s' if n != 1 else ''} found")
        self._search_lbl.configure(
            text=f"{n} result{'s' if n != 1 else ''}", fg="#16a34a")
        self._btn_search.configure(state="normal", text="  Search WebFIRE  ")
        self._set_status(f"Search complete — {n} reports found")

        # M9 — show/hide empty placeholder
        if n == 0:
            self._res_empty_lbl.pack(pady=12)
        else:
            self._res_empty_lbl.pack_forget()

        self._show_card(self._results_card)
        self._show_card(self._dl_card)
        self._show_card(self._scan_card)

        # Reset results sort state
        self._res_sort_col = None
        self._res_sort_asc = True

    def _on_search_error(self, msg):
        self._stop_search_spinner()

        self._session = None   # force session rebuild next time
        self._btn_search.configure(state="normal", text="  Search WebFIRE  ")
        self._search_lbl.configure(text="Search failed", fg="#dc2626")
        self._set_status(f"Error: {msg}")
        messagebox.showerror("Search Failed", msg)

    # ── Results sort (H7) ──────────────────────────────────────────────────

    def _on_res_sort(self, col: str):
        if col == "chk":
            return  # checkbox column — not sortable
        if self._res_sort_col == col:
            self._res_sort_asc = not self._res_sort_asc
        else:
            self._res_sort_col = col
            self._res_sort_asc = True

        _COL_FIELD = {
            "facility":  "facility",
            "city":      "city",     "st":   "state",
            "date":      "date",     "type": "report_type",
            "subtype":   "report_subtype", "pollutants": "pollutants",
            "status":    "downloaded",
        }
        field = _COL_FIELD.get(col, col)
        sorted_reports = sorted(
            self._reports,
            key=lambda r: str(r.get(field, "")).lower(),
            reverse=not self._res_sort_asc)

        self._res_tree.delete(*self._res_tree.get_children())
        for r in sorted_reports:
            cached  = r.get("downloaded", False)
            checked = self._selected.get(r["id"], False)
            self._insert_result_row(r, checked=checked, cached=cached)

        self._update_res_sort_heading()

    # ── Download ───────────────────────────────────────────────────────────

    def _do_download(self):
        ids = [iid for iid, sel in self._selected.items() if sel]
        if not ids:
            messagebox.showinfo("Nothing selected",
                                "Check at least one report to download.")
            return

        by_id   = {r["id"]: r for r in self._reports}
        targets = [by_id[i] for i in ids if i in by_id]
        if not targets:
            return

        dl_dir = self._active_folder
        dl_dir.mkdir(parents=True, exist_ok=True)
        self._expand_dl()

        self._dl_cancel = False
        self._btn_download.configure(state="disabled")
        self._btn_cancel_dl.pack(side="left", padx=(0, 4))
        self._dl_bar["maximum"] = len(targets)
        self._dl_bar["value"]   = 0
        self._dl_lbl.configure(text=f"0 / {len(targets)}")
        self._dl_file_lbl.configure(text="")
        self._dl_log.configure(state="normal")
        self._dl_log.delete("1.0", "end")
        self._dl_log.configure(state="disabled")
        self._set_status("Downloading reports…")

        def worker():
            session = self._session
            session.headers["Referer"] = core.BASE_URL + "eSearchResults.cfm"
            for rpt in targets:
                if self._dl_cancel:
                    break
                # M13 — show current file name
                self.after(0, lambda r=rpt:
                           self._dl_file_lbl.configure(
                               text=f"Downloading: {r['id']}  —  {r['facility'][:60]}"))
                ok, path, msg = core.download_report(
                    session, rpt["id"], dl_dir)
                label = "cached" if msg == "already cached" \
                    else ("ok" if ok else f"error: {msg}")
                icon  = "\u2713" if (ok and msg != "already cached") \
                    else ("\u21a9" if msg == "already cached" else "\u2717")
                self.after(0, lambda r=rpt, lbl=label, ic=icon:
                           self._on_dl_progress(r, lbl, ic))
                if ok and msg != "already cached":
                    sidecar = dl_dir / f"{rpt['id']}.json"
                    sidecar.write_text(json.dumps({
                        "facility":       rpt.get("facility", ""),
                        "city":           rpt.get("city", ""),
                        "state":          rpt.get("state", ""),
                        "date":           str(rpt.get("date", "")),
                        "report_type":    rpt.get("report_type", ""),
                        "organization":   rpt.get("organization", ""),
                        "county":         rpt.get("county", ""),
                        "report_subtype": rpt.get("report_subtype", ""),
                        "pollutants":     rpt.get("pollutants", ""),
                    }), encoding="utf-8")
                    time.sleep(0.8)
            self.after(0, self._on_dl_done)

        threading.Thread(target=worker, daemon=True).start()

    def _on_dl_progress(self, rpt, label, icon):
        done  = int(self._dl_bar["value"]) + 1
        total = int(self._dl_bar["maximum"])
        self._dl_bar["value"] = done
        self._dl_lbl.configure(text=f"{done} / {total}")

        line = f"{icon}  {rpt['id']}  —  {rpt['facility']}  ({label})\n"
        self._dl_log.configure(state="normal")
        self._dl_log.insert("end", line)
        self._dl_log.see("end")
        self._dl_log.configure(state="disabled")

        # Update status cell in results tree
        if rpt["id"] in self._res_tree.get_children():
            vals = list(self._res_tree.item(rpt["id"], "values"))
            if label in ("ok", "cached"):
                vals[8] = "Downloaded"
                self._res_tree.item(rpt["id"], values=vals, tags=("cached",))
            elif label.startswith("error"):
                vals[8] = "Error"
                self._res_tree.item(rpt["id"], values=vals)

    def _cancel_download(self):
        self._dl_cancel = True
        self._btn_cancel_dl.configure(state="disabled")
        # M12 — clearer cancelling message
        self._set_status("Cancelling — finishing current file…")

    def _on_dl_done(self):
        self._btn_cancel_dl.pack_forget()
        self._btn_cancel_dl.configure(state="normal")
        self._btn_download.configure(state="normal")
        self._dl_file_lbl.configure(text="")
        done  = int(self._dl_bar["value"])
        total = int(self._dl_bar["maximum"])
        if self._dl_cancel and done < total:
            self._set_status(
                f"Download cancelled — {done} of {total} completed")
        else:
            self._set_status(
                f"Download complete — {total} report{'s' if total != 1 else ''}")
        # H16 — refresh scan button now that files may exist
        self._refresh_scan_button_state()
        self._expand_scan()

    # ── Scan ───────────────────────────────────────────────────────────────

    def _meta_for_zip(self, z: Path) -> dict:
        """Return metadata for a ZIP, preferring the sidecar JSON written at download time."""
        sidecar = z.with_suffix(".json")
        if sidecar.exists():
            try:
                saved = json.loads(sidecar.read_text(encoding="utf-8"))
                return {
                    "id":             z.stem,
                    "facility":       saved.get("facility", "") or z.stem,
                    "city":           saved.get("city", ""),
                    "state":          saved.get("state", ""),
                    "date":           saved.get("date", ""),
                    "report_type":    saved.get("report_type", "") or core.classify_report(z),
                    "organization":   saved.get("organization", ""),
                    "county":         saved.get("county", ""),
                    "report_subtype": saved.get("report_subtype", ""),
                    "pollutants":     saved.get("pollutants", ""),
                    "downloaded":     True,
                    "scanned":        False,
                    "_zip_path":      str(z),
                }
            except Exception:
                pass
        rtype     = core.classify_report(z)
        file_meta = core.extract_file_meta(z)
        return {
            "id":             z.stem,
            "facility":       file_meta["facility"] or z.stem,
            "city":           file_meta["city"],
            "state":          file_meta["state"],
            "date":           "",
            "report_type":    rtype,
            "organization":   "",
            "county":         "",
            "report_subtype": "",
            "pollutants":     "",
            "downloaded":     True,
            "scanned":        False,
            "_zip_path":      str(z),
        }

    def _do_scan(self):
        if self._scan_rows:
            if not messagebox.askyesno("Rescan?",
                    f"This will replace {len(self._scan_rows)} existing findings.\nContinue?"):
                return

        scan_dir = self._active_folder
        zips_in_dir = sorted(scan_dir.glob("*.zip"))
        targets = []
        known = {r["id"]: r for r in self._reports}
        for z in zips_in_dir:
            rid = z.stem
            targets.append(known[rid] if rid in known else self._meta_for_zip(z))

        if not targets:
            messagebox.showinfo("Nothing to scan",
                                "Download at least one report first.")
            return

        self._disable_scan_buttons()
        self._scan_zip_dir = scan_dir
        self._start_scan_ui(len(targets))
        self._set_status("Scanning reports…")

        def worker():
            for rpt in targets:
                path = scan_dir / f"{rpt['id']}.zip"
                rows = core.scan_report(path, rpt)
                self.after(0, lambda r=rows, total=len(targets):
                           self._on_scan_progress(r, total))
            self.after(0, self._on_scan_done)

        threading.Thread(target=worker, daemon=True).start()

    def _on_scan_progress(self, new_rows, total):
        self._scan_rows.extend(new_rows)
        done = int(self._scan_bar["value"]) + 1
        self._scan_bar["value"] = done
        self._scan_lbl.configure(text=f"{done} / {total}")

        if new_rows:
            self._insert_report_group(new_rows)

        flagged     = sum(1 for r in self._scan_rows if r.get("deviation") == "YES")
        need_review = sum(1 for r in self._scan_rows if r.get("deviation") == "manual-review")
        n           = len(self._scan_rows)

    def _insert_report_group(self, rows: list, parent_iid: str = ""):
        if not rows:
            return
        first     = rows[0]
        report_id = str(first.get("report_id", id(rows)))

        if self._scan_tree.exists(report_id):
            report_id = f"{report_id}_{len(self._scan_tree.get_children())}"

        deviations = [r for r in rows if r.get("deviation") == "YES"]
        reviews    = [r for r in rows if r.get("deviation") == "manual-review"]
        errors     = [r for r in rows if r.get("deviation") == "error"]
        n_pass     = len(rows) - len(deviations) - len(reviews) - len(errors)

        if deviations:
            agg_result, agg_tag = "DEVIATION",     "deviation"
        elif reviews:
            agg_result, agg_tag = "Manual Review", "manualreview"  # M17
        elif errors:
            agg_result, agg_tag = "Error",         "error"
        else:
            agg_result, agg_tag = "Pass",          ""

        parts = []
        if deviations: parts.append(f"{len(deviations)} DEVIATION")
        if reviews:    parts.append(f"{len(reviews)} Manual Review")   # M17
        if errors:     parts.append(f"{len(errors)} Error")
        if n_pass:     parts.append(f"{n_pass} Pass")
        parent_notes = (f"{len(rows)} findings — " + ", ".join(parts)) if len(rows) > 1 else ""

        auto_open = bool(deviations or reviews)
        self._scan_tree.insert(
            parent_iid, "end", iid=report_id, open=auto_open,
            values=(first.get("report_id", ""),
                    agg_result,
                    first.get("facility", ""),
                    first.get("state", ""),
                    str(first.get("date", ""))[:10],
                    first.get("report_type", ""),
                    "", "", "", "", "",
                    parent_notes),
            tags=(agg_tag,) if agg_tag else ())

        for idx, row in enumerate(rows):
            child_iid = f"{report_id}_c{idx}"
            dev       = row.get("deviation", "")
            is_aer    = row.get("report_type", "") == "AER"

            if dev == "error":          child_tag = "error"
            elif dev == "YES":          child_tag = "deviation"
            elif dev == "manual-review":child_tag = "manualreview"
            elif dev == "no limit":     child_tag = "manualreview"
            else:                       child_tag = ""

            result_label = {
                "YES":           "DEVIATION",
                "no":            "Pass",
                "no limit":      "Manual Review",
                "error":         "Error",
                "manual-review": "Manual Review",  # M17
                "count-only":    "Count Only",
            }.get(dev, dev)

            if is_aer:
                location_val  = row.get("sheet",       row.get("location", ""))
                pollutant_val = row.get("citation",     row.get("pollutant", ""))
                measured_val  = ""
                limit_val     = "—"
                pct_str       = "—"
                notes_val     = row.get("description", row.get("error", ""))
            else:
                location_val  = row.get("location", "")
                pollutant_val = row.get("pollutant", "")
                measured_val  = row.get("avg_measured", "")
                pct           = row.get("pct_of_limit", "")
                pct_str       = f"{pct}%" if pct != "" else "—"
                limit_val     = row.get("limit", "") or "—"
                notes_val     = row.get("error", "")

            self._scan_tree.insert(
                report_id, "end", iid=child_iid,
                values=("",
                        result_label,
                        "",
                        "",
                        "",
                        "",
                        location_val,
                        pollutant_val,
                        measured_val,
                        limit_val,
                        pct_str,
                        notes_val),
                tags=(child_tag,) if child_tag else ())

    # M18 — double-click opens the actual report document
    def _open_scan_result_folder(self, event=None):
        iid = self._scan_tree.focus()
        if not iid:
            return

        # Derive report_id: parent rows use report_id as iid; child rows use "{id}_cN"
        parent = self._scan_tree.parent(iid)
        report_id = parent if parent else iid

        if self._scan_zip_dir is None:
            self._open_path(self._active_folder)
            return

        zip_path = self._scan_zip_dir / f"{report_id}.zip"
        if not zip_path.exists():
            self._open_path(self._active_folder)
            return

        # Extract to a per-report temp directory and open the main document
        import tempfile
        tmp = Path(tempfile.mkdtemp(prefix=f"cdx_{report_id}_"))
        try:
            with zipfile.ZipFile(zip_path) as zf:
                zf.extractall(tmp)
            all_files = [f for f in tmp.rglob("*") if f.is_file()]
            to_open = [f for f in all_files if f.suffix.lower() in (".pdf", ".xlsx", ".xlsm")]
            if not to_open:
                # Fallback to XML if no PDF/spreadsheet found
                to_open = [f for f in all_files if f.suffix.lower() == ".xml"]
            if to_open:
                for f in to_open:
                    self._open_path(f)
                self.after(10_000, lambda t=tmp: shutil.rmtree(t, ignore_errors=True))
                return
            # Fallback: open the extracted folder
            self._open_path(tmp)
            self.after(10_000, lambda t=tmp: shutil.rmtree(t, ignore_errors=True))
        except Exception:
            self._open_path(self._active_folder)

    # ── Scan filter / sort ─────────────────────────────────────────────────

    def _clear_filter(self):
        self._filter_var.set("")
        self._filter_result_var.set("All")
        self._apply_scan_filter()

    def _on_scan_sort(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._apply_scan_filter()
        self._update_scan_sort_heading()  # H15

    def _apply_scan_filter(self):
        if not self._scan_rows:
            return

        filter_text   = (self._filter_var.get() if self._filter_var else "").lower().strip()
        filter_result = self._filter_result_var.get() if self._filter_result_var else "All"

        # M17 — "Manual Review" maps to "manual-review"
        _RESULT_DEV_MAP = {
            "Deviations":    "YES",
            "Manual Review": "manual-review",
            "Errors":        "error",
        }
        filter_dev = _RESULT_DEV_MAP.get(filter_result)

        def _row_text(r):
            return " ".join([
                str(r.get("facility",    "")),
                str(r.get("citation",    "")),
                str(r.get("pollutant",   "")),
                str(r.get("sheet",       "")),
                str(r.get("location",    "")),
                str(r.get("description", "")),
                str(r.get("error",       "")),
            ]).lower()

        groups: dict[str, list] = {}
        for row in self._scan_rows:
            rid = str(row.get("report_id", ""))
            groups.setdefault(rid, []).append(row)

        filtered_groups = []
        for rid, rows in groups.items():
            if filter_result == "Pass":
                matched = [r for r in rows if r.get("deviation") == "no"]
            elif filter_dev:
                matched = [r for r in rows if r.get("deviation") == filter_dev]
            else:
                matched = rows[:]

            if filter_text:
                facility_match = filter_text in str(rows[0].get("facility", "")).lower()
                matched = [r for r in matched
                           if facility_match or filter_text in _row_text(r)]

            if matched:
                filtered_groups.append((rid, matched))

        _COL_KEY = {
            "frs_id":   lambda g: str(g[0]).lower(),
            "result":   lambda g: _agg_result_rank(g[1]),
            "facility": lambda g: str(g[1][0].get("facility", "")).lower(),
            "st":       lambda g: str(g[1][0].get("state", "")).lower(),
            "date":     lambda g: str(g[1][0].get("date", "")),
            "type":     lambda g: str(g[1][0].get("report_type", "")).lower(),
            "pollutant":lambda g: str(g[1][0].get("citation",
                                       g[1][0].get("pollutant", ""))).lower(),
            "notes":    lambda g: str(g[1][0].get("description",
                                       g[1][0].get("error", ""))).lower(),
        }
        sort_key = _COL_KEY.get(self._sort_col, _COL_KEY["date"])
        filtered_groups.sort(key=sort_key, reverse=not self._sort_asc)

        self._scan_tree.delete(*self._scan_tree.get_children())
        for rid, rows in filtered_groups:
            self._insert_report_group(rows)

        # L24 — fallback as full-width strip
        n_fallback = sum(1 for r in self._scan_rows if r.get("fallback_used"))
        if n_fallback:
            unmatched = list({r.get("unmatched_citation", "")
                              for r in self._scan_rows if r.get("fallback_used")
                              and r.get("unmatched_citation")})
            self._fallback_lbl.configure(
                text=f"\u26a0  {n_fallback} finding(s) used keyword fallback routing"
                     + (f"  —  citations: {', '.join(unmatched)}" if unmatched else ""),
                pady=4)
        else:
            self._fallback_lbl.configure(text="", pady=0)

    def _on_scan_done(self):
        self._scan_prog_frame.pack_forget()
        flagged     = sum(1 for r in self._scan_rows if r.get("deviation") == "YES")
        need_review = sum(1 for r in self._scan_rows if r.get("deviation") == "manual-review")
        n           = len(self._scan_rows)

        no_issues = n - flagged - need_review
        msg = (f"Potential Deviations: {flagged}"
               f"  |  Manual Review: {need_review}"
               f"  |  No Issues Identified: {no_issues}")
        if flagged:
            color = C_DEV_FG
        elif need_review:
            color = "#92400e"
        else:
            color = "#15803d"

        self._scan_summary.configure(text=msg, fg=color)
        self._scan_summary.pack(fill="x", before=self._tree_frame)
        self._btn_scan.configure(command=self._do_scan, cursor="hand2")
        if self._scan_rows:
            self._btn_export.configure(state="normal")
            self._btn_export_xlsx.configure(state="normal")
            self._btn_extract.configure(state="normal")

        # M26 — update window title with summary
        self.title(
            f"WebFIRE Deviation Scanner  —  {n} findings"
            + (f"  |  {flagged} deviation{'s' if flagged != 1 else ''}" if flagged else ""))

        self._set_status(f"Scan complete — {n} findings, {flagged} flagged")
        self._apply_scan_filter()
        self._scan_tree.focus_set()

    # ── Export helpers ─────────────────────────────────────────────────────

    def _get_filtered_rows(self):
        """Return rows matching the current filter state, preserving sort order."""
        filter_result = self._filter_result_var.get() if self._filter_result_var else "All"
        filter_text   = (self._filter_var.get() if self._filter_var else "").lower().strip()
        _RESULT_DEV_MAP = {
            "Deviations":    "YES",
            "Manual Review": "manual-review",
            "Errors":        "error",
        }
        filter_dev = _RESULT_DEV_MAP.get(filter_result)

        def _row_text(r):
            return " ".join([
                str(r.get("facility",    "")),
                str(r.get("citation",    "")),
                str(r.get("pollutant",   "")),
                str(r.get("sheet",       "")),
                str(r.get("location",    "")),
                str(r.get("description", "")),
                str(r.get("error",       "")),
            ]).lower()

        result = []
        groups: dict[str, list] = {}
        for row in self._scan_rows:
            rid = str(row.get("report_id", ""))
            groups.setdefault(rid, []).append(row)

        for rid, rows in groups.items():
            if filter_result == "Pass":
                matched = [r for r in rows if r.get("deviation") == "no"]
            elif filter_dev:
                matched = [r for r in rows if r.get("deviation") == filter_dev]
            else:
                matched = rows[:]
            if filter_text:
                facility_match = filter_text in str(rows[0].get("facility", "")).lower()
                matched = [r for r in matched
                           if facility_match or filter_text in _row_text(r)]
            result.extend(matched)
        return result

    # ── Export CSV ─────────────────────────────────────────────────────────

    def _do_export(self):
        if not self._scan_rows:
            return
        rows_to_export = self._get_filtered_rows()
        filter_result  = self._filter_result_var.get() if self._filter_result_var else "All"
        default_name   = (f"{filter_result.lower().replace(' ', '_')}_filtered.csv"
                          if filter_result != "All" else "deviations.csv")
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=default_name,
            title="Save scan results")
        if not path:
            return
        fields = [
            "report_id", "facility", "city", "state", "date",
            "report_type", "deviation", "error",
            "location", "pollutant", "unit",
            "n_runs", "avg_measured", "limit", "pct_of_limit", "regulation",
            "sheet", "description",
        ]
        # Rename report_id → frs_id in the CSV header
        header_labels = ["frs_id"] + fields[1:]
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            f.write(",".join(header_labels) + "\r\n")
            writer.writerows(rows_to_export)
        self._set_status(f"Exported {len(rows_to_export)} rows → {Path(path).name}")

    # ── Export XLSX ────────────────────────────────────────────────────────

    def _do_export_xlsx(self):
        if not self._scan_rows:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for XLSX export.\n\nRun:  pip install openpyxl")
            return

        filter_result  = self._filter_result_var.get() if self._filter_result_var else "All"
        default_name   = (f"{filter_result.lower().replace(' ', '_')}_filtered.xlsx"
                          if filter_result != "All" else "deviations.xlsx")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_name,
            title="Save scan results as Excel workbook")
        if not path:
            return

        _IDENTITY = [
            ("report_id",    "Report ID",   14),
            ("facility",     "Facility",    28),
            ("city",         "City",        16),
            ("state",        "State",        6),
            ("date",         "Date",        12),
            ("deviation",    "Deviation",   12),
        ]
        ST_COLS = _IDENTITY + [
            ("location",     "Location",   20),
            ("pollutant",    "Pollutant",  24),
            ("unit",         "Unit",       18),
            ("n_runs",       "Runs",        6),
            ("avg_measured", "Measured",   12),
            ("limit",        "Limit",      12),
            ("pct_of_limit", "% of Limit", 10),
            ("regulation",   "Regulation", 20),
            ("error",        "Notes",      36),
        ]
        AER_COLS = _IDENTITY + [
            ("citation",    "Citation",    20),
            ("sheet",       "Sheet",       24),
            ("description", "Description", 50),
        ]
        ALL_COLS = [
            ("report_id",    "Report ID",            14),
            ("facility",     "Facility",              28),
            ("city",         "City",                  16),
            ("state",        "State",                  6),
            ("date",         "Date",                  12),
            ("report_type",  "Type",                   6),
            ("deviation",    "Deviation",             12),
            ("error",        "Notes / Description",   40),
            ("location",     "Location / Sheet",      22),
            ("pollutant",    "Pollutant / Citation",  24),
            ("unit",         "Unit",                  18),
            ("n_runs",       "Runs",                   6),
            ("avg_measured", "Measured",              12),
            ("limit",        "Limit",                 12),
            ("pct_of_limit", "% of Limit",            10),
            ("regulation",   "Regulation",            20),
            ("sheet",        "Sheet",                 22),
            ("description",  "Description",           40),
        ]

        filtered_rows = self._get_filtered_rows()
        if filter_result != "All":
            tab_defs = [
                (filter_result, filtered_rows, ALL_COLS),
            ]
        else:
            tab_defs = [
                ("Deviations",    [r for r in self._scan_rows if r.get("deviation") == "YES"],            ALL_COLS),
                ("Manual Review", [r for r in self._scan_rows if r.get("deviation") == "manual-review"],  ALL_COLS),  # M17
                ("Errors",        [r for r in self._scan_rows if r.get("deviation") == "error"],           ALL_COLS),
                ("Pass",          [r for r in self._scan_rows if r.get("deviation") == "no"],              ALL_COLS),
                ("All Results",   self._scan_rows,                                                         ALL_COLS),
            ]

        HDR_FILL = PatternFill("solid", fgColor="1A3A5C")
        HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
        DEV_FILL = PatternFill("solid", fgColor="FFE0E0")
        REV_FILL = PatternFill("solid", fgColor="FFF7E6")
        ERR_FILL = PatternFill("solid", fgColor="FFFBE6")

        _FILL_MAP = {
            "YES":           DEV_FILL,
            "manual-review": REV_FILL,
            "error":         ERR_FILL,
        }

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for tab_name, rows, col_defs in tab_defs:
            ws = wb.create_sheet(title=tab_name)

            for col_i, (field, label, width) in enumerate(col_defs, start=1):
                cell = ws.cell(row=1, column=col_i, value=label)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[
                    ws.cell(row=1, column=col_i).column_letter
                ].width = width

            for row_i, row in enumerate(rows, start=2):
                fill = _FILL_MAP.get(row.get("deviation", ""))
                for col_i, (field, label, _) in enumerate(col_defs, start=1):
                    val = row.get(field, "")
                    if isinstance(val, bool):
                        val = "Yes" if val else ""
                    cell = ws.cell(row=row_i, column=col_i, value=val)
                    if fill:
                        cell.fill = fill

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        try:
            wb.save(path)
            self._set_status(f"Exported {len(filtered_rows)} rows → {Path(path).name}")
        except PermissionError:
            messagebox.showerror(
                "Save Failed",
                f"Cannot write to {path}\n\nIs the file open in Excel?")

    # ── Extract Files ──────────────────────────────────────────────────────

    def _do_extract_files(self):
        if not self._scan_rows:
            return
        if self._scan_zip_dir is None:
            messagebox.showerror("Extract Error",
                                 "No scan directory known. Run a scan first.")
            return

        dest = filedialog.askdirectory(
            title="Choose destination folder for extracted files",
            initialdir=str(Path.home()))
        if not dest:
            return
        dest = Path(dest)

        # Group findings by report_id
        groups: dict = collections.defaultdict(list)
        for row in self._scan_rows:
            groups[str(row.get("report_id", ""))].append(row)

        # Build extraction plan
        plan = []
        for rid, rows in groups.items():
            zip_path = self._scan_zip_dir / f"{rid}.zip"
            subfolder = _aggregate_report_status(rows)
            plan.append((rid, zip_path, subfolder))

        if not plan:
            return

        # Create the three destination subfolders
        for name in ("Deviations", "Manual Review", "No Deviations"):
            (dest / name).mkdir(parents=True, exist_ok=True)

        # Progress dialog
        dialog = tk.Toplevel(self)
        dialog.title("Extracting Files")
        dialog.geometry("380x130")
        dialog.resizable(False, False)
        dialog.configure(bg="white")
        dialog.transient(self)
        dialog.grab_set()

        tk.Frame(dialog, bg=C_HEADER, height=4).pack(fill="x")
        progress_lbl = tk.Label(dialog, text=f"Processing 0 / {len(plan)}…",
                                bg="white", fg="#374151",
                                font=("Helvetica", 10))
        progress_lbl.pack(pady=(14, 6))
        progress_bar = ttk.Progressbar(dialog, mode="determinate",
                                       maximum=len(plan), length=320)
        progress_bar.pack()

        cancel_event = threading.Event()

        def _on_cancel():
            cancel_event.set()
            dialog.destroy()

        ttk.Button(dialog, text="Cancel", style="Grey.TButton",
                   cursor="hand2", command=_on_cancel).pack(pady=10)

        errors = []
        extracted_counts = {"Deviations": 0, "Manual Review": 0, "No Deviations": 0}

        def _update_progress(done):
            if not cancel_event.is_set() and dialog.winfo_exists():
                progress_bar["value"] = done
                progress_lbl.configure(text=f"Processing {done} / {len(plan)}…")

        def _on_done():
            if dialog.winfo_exists():
                dialog.destroy()
            total = sum(extracted_counts.values())
            lines = [
                f"Extraction complete — {total} of {len(plan)} reports extracted.",
                f"  Deviations:    {extracted_counts['Deviations']} report(s)",
                f"  Manual Review: {extracted_counts['Manual Review']} report(s)",
                f"  No Deviations: {extracted_counts['No Deviations']} report(s)",
            ]
            if errors:
                lines.append(f"\n{len(errors)} error(s):")
                lines.extend(f"  {e}" for e in errors[:10])
                if len(errors) > 10:
                    lines.append(f"  …and {len(errors) - 10} more")
            messagebox.showinfo("Extraction Complete", "\n".join(lines))
            self._set_status(f"Extracted {total} reports → {dest.name}/")
            if messagebox.askyesno("Open Folder?",
                                   f"Open destination folder in Finder/Explorer?\n{dest}"):
                self._open_path(dest)

        def worker():
            for i, (rid, zip_path, subfolder) in enumerate(plan):
                if cancel_event.is_set():
                    break
                target_dir = dest / subfolder
                if not zip_path.exists():
                    errors.append(f"{rid}: ZIP not found at {zip_path}")
                    self.after(0, lambda i=i: _update_progress(i + 1))
                    continue
                try:
                    with zipfile.ZipFile(zip_path, "r") as zf:
                        for member in zf.namelist():
                            original_name = Path(member).name
                            if not original_name:   # skip directory entries
                                continue
                            if original_name.lower() == "metadata.xml":
                                continue
                            out_path = target_dir / f"{rid}_{original_name}"
                            out_path.write_bytes(zf.read(member))
                    extracted_counts[subfolder] += 1
                except zipfile.BadZipFile:
                    # WebFIRE returned a bare PDF saved under the .zip name.
                    # Copy it directly to the target folder as a .pdf file.
                    with open(zip_path, "rb") as f:
                        raw = f.read(4)
                    if raw == b"%PDF":
                        out_path = target_dir / f"{rid}.pdf"
                        out_path.write_bytes(zip_path.read_bytes())
                        extracted_counts[subfolder] += 1
                    else:
                        errors.append(f"{rid}: corrupt or invalid ZIP")
                except Exception as exc:
                    errors.append(f"{rid}: {exc}")
                self.after(0, lambda i=i: _update_progress(i + 1))
            self.after(0, _on_done)

        threading.Thread(target=worker, daemon=True).start()

    # ── Shared UI helpers ──────────────────────────────────────────────────

    def _open_path(self, path: Path):
        """Open a file or folder in the system default application."""
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception:
            pass

    def _bind_scroll_suppression(self, widget):
        widget.bind("<Enter>", lambda _: self._canvas.unbind("<MouseWheel>"))
        widget.bind("<Leave>", lambda _: self._canvas.bind("<MouseWheel>", self._on_mousewheel))

    def _stop_search_spinner(self):
        self._search_spinner.stop()
        self._search_spinner.pack_forget()

    def _disable_scan_buttons(self):
        self._btn_scan.configure(command=lambda: None, cursor="arrow")
        for btn in (self._btn_export, self._btn_export_xlsx, self._btn_extract):
            btn.configure(state="disabled")

    def _start_scan_ui(self, n_targets: int):
        """Reset and show scan progress UI; caller sets _scan_zip_dir and status."""
        self._scan_rows = []
        self._scan_tree.delete(*self._scan_tree.get_children())
        self._scan_summary.pack_forget()
        self._scan_prog_frame.pack(fill="x", before=self._tree_frame)
        self._scan_bar["maximum"] = n_targets
        self._scan_bar["value"]   = 0
        self._scan_lbl.configure(text=f"0 / {n_targets}")

    def _insert_result_row(self, r: dict, checked: bool, cached: bool):
        chk = "\u2611" if checked else "\u2610"
        self._res_tree.insert(
            "", "end", iid=r["id"],
            values=(chk,
                    r["facility"],
                    r["city"], r["state"],
                    r["date"][:10],
                    r["report_type"], r["report_subtype"],
                    r["pollutants"][:60],
                    "Downloaded" if cached else "Pending"),
            tags=("cached",) if cached else ())

    # ── Help / About (M27) ─────────────────────────────────────────────────

    def _show_help(self):
        win = tk.Toplevel(self)
        win.title("About WebFIRE Deviation Scanner")
        win.geometry("480x340")
        win.resizable(False, False)
        win.configure(bg="white")
        win.transient(self)
        win.grab_set()

        tk.Frame(win, bg=C_HEADER, height=6).pack(fill="x")

        tk.Label(win, text="WebFIRE Deviation Scanner",
                 bg="white", fg=C_HEADER,
                 font=("Helvetica", 14, "bold")).pack(pady=(18, 4))
        tk.Label(win, text=f"Version {VERSION}",
                 bg="white", fg="#64748b",
                 font=("Helvetica", 10)).pack()

        sep = tk.Frame(win, bg="#e2e8f0", height=1)
        sep.pack(fill="x", padx=20, pady=14)

        body_text = (
            "This tool searches EPA WebFIRE for stack test and CEDRI compliance\n"
            "reports, downloads them, and automatically scans for deviations\n"
            "from regulatory emission limits.\n\n"
            "Workflow:\n"
            "  1. Search WebFIRE using facility name, state, CFR part, or dates.\n"
            "  2. Select and download the reports you want to review.\n"
            "  3. Click Scan for Deviations to analyze all downloaded files.\n"
            "  4. Export results to CSV or XLSX for further analysis.\n\n"
            "Tip: Use Change Folder to point the app at any folder of ZIPs,\n"
            "then Scan for Deviations \u2014 no re-search needed."
        )
        tk.Label(win, text=body_text, bg="white", fg="#374151",
                 font=("Helvetica", 10), justify="left", anchor="w",
                 padx=24).pack(fill="x")

        ttk.Button(win, text="Close", style="Blue.TButton",
                   cursor="hand2",
                   command=win.destroy).pack(pady=(16, 0))


# ── Helpers ────────────────────────────────────────────────────────────────

def _aggregate_report_status(rows: list) -> str:
    """Return extraction subfolder name for a group of findings from one report.
    Priority: YES > manual-review > error > no/count-only
    """
    devs = {r.get("deviation", "") for r in rows}
    if "YES"           in devs: return "Deviations"
    if "manual-review" in devs: return "Manual Review"
    if "error"         in devs: return "Manual Review"
    return "No Deviations"


def _agg_result_rank(rows: list) -> int:
    devs = {r.get("deviation", "") for r in rows}
    if "YES"           in devs: return 0
    if "error"         in devs: return 1
    if "manual-review" in devs: return 2
    return 3


def _cfr_value(display: str) -> str:
    for val, label in core.CFR_PART_OPTIONS:
        if label == display or val == display:
            return val
    return "All"


# ── Entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
