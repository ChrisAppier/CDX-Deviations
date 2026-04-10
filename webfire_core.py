"""
webfire_core.py
---------------
Core search / download / scan logic. Imported by gui.py.

Supports two report categories:
  ST  — Stack Test (XML-based, existing logic)
  AER — Annual/Semiannual Emissions Reports (Excel .xlsx/.xlsm, CEDRI templates)
"""

import io
import warnings
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

import requests
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore")  # suppress LibreSSL urllib3 noise

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_URL    = "https://cfpub.epa.gov/webfire/reports/"
DETAIL_URL  = "https://cfpub.epa.gov/webfire/FIRE/view/dspERTDocumentDetails.cfm"
_UA         = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/16.0 Safari/605.1.15"
)

STATE_OPTIONS = [
    ("AA", "All States"), ("AL", "ALABAMA"), ("AK", "ALASKA"), ("AS", "AMERICAN SAMOA"),
    ("AZ", "ARIZONA"), ("AR", "ARKANSAS"), ("CA", "CALIFORNIA"), ("CO", "COLORADO"),
    ("CT", "CONNECTICUT"), ("DE", "DELAWARE"), ("DC", "DISTRICT OF COLUMBIA"),
    ("FL", "FLORIDA"), ("GA", "GEORGIA"), ("GU", "GUAM"), ("HI", "HAWAII"),
    ("ID", "IDAHO"), ("IL", "ILLINOIS"), ("IN", "INDIANA"), ("IA", "IOWA"),
    ("KS", "KANSAS"), ("KY", "KENTUCKY"), ("LA", "LOUISIANA"), ("ME", "MAINE"),
    ("MD", "MARYLAND"), ("MA", "MASSACHUSETTS"), ("MI", "MICHIGAN"),
    ("MN", "MINNESOTA"), ("MS", "MISSISSIPPI"), ("MO", "MISSOURI"), ("MT", "MONTANA"),
    ("NE", "NEBRASKA"), ("NV", "NEVADA"), ("NH", "NEW HAMPSHIRE"),
    ("NJ", "NEW JERSEY"), ("NM", "NEW MEXICO"), ("NY", "NEW YORK"),
    ("NC", "NORTH CAROLINA"), ("ND", "NORTH DAKOTA"),
    ("MP", "NORTHERN MARIANA ISLANDS"), ("OH", "OHIO"), ("OK", "OKLAHOMA"),
    ("OR", "OREGON"), ("PA", "PENNSYLVANIA"), ("PR", "PUERTO RICO"),
    ("RI", "RHODE ISLAND"), ("SC", "SOUTH CAROLINA"), ("SD", "SOUTH DAKOTA"),
    ("TN", "TENNESSEE"), ("TX", "TEXAS"), ("UT", "UTAH"), ("VT", "VERMONT"),
    ("VI", "VIRGIN ISLANDS"), ("VA", "VIRGINIA"), ("WA", "WASHINGTON"),
    ("WV", "WEST VIRGINIA"), ("WI", "WISCONSIN"), ("WY", "WYOMING"),
]

CFR_PART_OPTIONS = [
    ("All",     "All"),
    ("Part 60", "Part 60 — NSPS"),
    ("Part 62", "Part 62 — Federal Plan"),
    ("Part 63", "Part 63 — NESHAP"),
]


# ---------------------------------------------------------------------------
# AER routing and detection constants
# ---------------------------------------------------------------------------

# Maps CitationID substrings to the sheet names that contain deviation data.
# Keys are matched as substrings of the CitationID (case-insensitive).
CITATION_TO_SHEETS = {
    "60.5420a": ["Deviations"],
    "60.395":   ["CMS_Deviation_Detail", "Emiss_Data_Detail"],
    "63.1354":  ["Excess Emissions", "Excess Emissions Summary",
                 "Malfunction Deviation Count", "CMS Downtime Detail"],
    "63.5912":  ["Summary_Report_Limits", "Summary_Report_CMS",
                 "Limits_Detail", "CMS_Detail"],
    "63.6150":  ["Deviation_Limits", "Deviation_Summary_Limits",
                 "Deviation_CEM_CPMS", "Deviation_Summary_CPMS",
                 "Number_of_Deviations"],
    "63.4520":  ["Number_of_Deviations_Compliant", "Number_of_Deviations_Without",
                 "Number_of_Deviations_Work", "With_Addon_Malfunction",
                 "With_Addon_OpLimit"],
    "63.6650":  ["Fuel Req. Deviation", "Voltage Frequency Deviation",
                 "Non CMS Deviation", "CMS Deviation Detail",
                 "CMS Deviation Summary", "Landfill Digester Gas Deviation",
                 "Malfunctions"],
    "63.5580":  ["No CMS - Deviation - Limits", "Using CMS - Deviation - Limits",
                 "Using CMS - Deviation - Summary", "Using CMS - Malfunction",
                 "Deviation - Daily Averages"],
    "63.5765":  ["Deviation_Limits", "Summary_Report_Limit", "Summary_Report_CPMS"],
    "63.10899": ["CMS_Deviation_Detail", "CMS_Deviation_Summary",
                 "Limit_Deviation_Detail", "Limit_Deviation_Summary"],
    "63.7751":  ["CMS_Deviation_Detail", "CMS_Deviation_Summary",
                 "Limit_Deviation_Detail", "Limit_Deviation_Summary"],
    "63.3400":  ["Deviations_Limit_No_CMS", "Deviations_Op_Param_CMS",
                 "OpParam_Deviation_CMS_Data", "Deviations_Op_Param_CMS_Summary"],
    "63.7550":  ["Deviations No CMS", "Deviations w CMS",
                 "Deviations w CMS Summary", "Malfunctions", "CMS Downtime"],
    "63.7951":  ["No CMS - Deviation", "Deviation - CMS", "Deviation - CMS Summary",
                 "Deviation - Limits", "Deviation - Limits Summary",
                 "Number of Deviations"],
    "63.3120":  ["Deviations_Work_Practice_Plans", "Malfunction",
                 "Number_of_Deviations"],
    "63.3920":  ["Number_of_Deviations", "Number_of_Deviations_WorkPracti",
                 "With_Addon_Malfunction"],
    # NSPS KKKK / Lead Acid Battery Mfg — matches multi-citation containing "63.11424"
    "63.11424":  ["CMS_Downtime", "CMS_Downtime_Summary", "CMS_Downtime_Summary_Only",
                  "Excess_Emissions_Details", "Excess_Emissions_Summary"],
    # NESHAP UUU — Rubber Tire Manufacturing (§63.6009/§63.6010)
    "63.6009":   ["Compliance Rpt - Deviations"],
    "63.4214":  [],   # handled separately (non-emergency hours check)
    "63.655":   [],   # fenceline — handled separately via _scan_fenceline_sheet
    "60.5422b": [],   # OOOOb — leak records require domain logic
    "60.5422a": [],   # OOOOa semiannual — flat Semiannual_Report sheet, no deviation table
}

# Keyword fragments used as fallback when no CitationID matches a routing entry.
# Any sheet whose name contains one of these (case-insensitive) is scanned.
KEYWORD_FALLBACK = ["deviat", "excess", "malfunction", "malfunc", "downtime"]

# Substrings in a field name that indicate it is a numeric deviation summary/count.
# A non-zero value in such a field means a deviation is present.
# Note: the actual column names in CEDRI templates use variations like
# DeviationsCount_Number, TotalDeviationHrs, CMSDowntimeDuration, Number_DNC, etc.
_SUMMARY_FIELD_KEYWORDS = (
    "duration", "hrs", "hours", "percent",
    "_dnc", "dnc_", "number_dnc", "number_of_dev",
    "deviationcount", "deviationscount", "downtime",
    "count_number", "_count", "count_no",   # count-based deviation fields
    "devs",          # NumOfDevs, WorkPracticeDevs, OrganicHAPDevs (63.3120 template)
    "numofdev",      # NumOfDeviations abbreviation patterns
)

# Column index / row positions in the standard CEDRI template.
# The CitationID label is in column A; the value is in column B (index 1).
# NOTE: EPA documentation says column C but empirical review of hundreds of
# actual files consistently shows the value in column B.
_CITATION_ROW = 2   # 0-indexed (row 3 in Excel)
_CITATION_COL = 1   # 0-indexed (column B in Excel — label is col A, value is col B)

# Domain-specific thresholds
FENCELINE_THRESHOLD_UGM3 = 9.0      # §63.655: benzene action level (µg/m³ annual avg)
TURBINE_ANNUAL_LIMIT_HRS = 100.0    # §60.4214: non-emergency use limit per year

# Maps lowercase field names → human-readable phrase template.
# {v} is the numeric value.  Used by _format_summary_description().
_FIELD_FRIENDLY = {
    "totaldeviationhrs":             "{v:.3g} total deviation hr(s)",
    "totaldeviationduration":        "{v:.3g} total deviation hr(s)",
    "emissionsdeviationduration":    "{v:.3g} hr(s) of emission exceedance",
    "emissionsdeviationpercent":     "{v:.3g}% of operating time with exceedance",
    "cmsdowntimeduration":           "{v:.3g} hr(s) of CMS downtime",
    "cmsdowntimepercent":            "{v:.3g}% CMS downtime",
    "cmsoutageduration":             "{v:.3g} hr(s) of CMS outage",
    "deviationduration":             "{v:.3g} hr(s) of deviation",
    "deviationcountstandard":        "{v:.0f} standard deviation(s)",
    "deviationcountdowntime":        "{v:.0f} monitoring downtime period(s)",
    "cpmsoutageduration":            "{v:.3g} hr(s) of CPMS outage",
    "cpmsdowntimeduration":          "{v:.3g} hr(s) of CPMS downtime",
    "cpmsdowntimepercent":           "{v:.3g}% CPMS downtime",
    "excessduration":                "{v:.3g} hr(s) of excess emissions",
    "parameterexcessduration":       "{v:.3g} hr(s) of excess emissions",
    "downtimeduration":              "{v:.3g} hr(s) of CMS downtime",
    "number_dnc":                    "{v:.0f} deviation(s) without CMS",
    "othercausehrs":                 "{v:.3g} hr(s) — other cause",
    "unknowncausehrs":               "{v:.3g} hr(s) — unknown cause",
    "malfunctionhrs":                "{v:.3g} hr(s) — malfunction",
    "totaldeviationpercent":         "{v:.3g}% of period with deviation",
    "numofdevs":                     "{v:.0f} deviation(s)",
    "workpracticedevs":              "{v:.0f} work practice deviation(s)",
    "organichapdevs":                "{v:.0f} organic HAP deviation(s)",
    "deviationscount_number":        "{v:.0f} deviation(s)",
    # CPMS/CMS downtime breakdown fields (63.6150 template)
    "cpmsdowntimemonitoringmal":     "{v:.0f} monitoring malfunction(s)",
    "cpmsdowntimecalibrations":      "{v:.0f} calibration(s)",
    "cpmsdowntimequality":           "{v:.0f} quality assurance event(s)",
    "cpmsdowntimeotherknown":        "{v:.0f} other known downtime period(s)",
    "cpmsdowntimeunknown":           "{v:.0f} unknown downtime period(s)",
    # Equipment/process problem hours (63.1354 template)
    "equipproblemhrs":               "{v:.3g} hr(s) — equipment problem",
    "processproblhrs":               "{v:.3g} hr(s) — process problem",
    "nonequipproblhrs":              "{v:.3g} hr(s) — non-equipment cause",
}


# ---------------------------------------------------------------------------
# Session
# ---------------------------------------------------------------------------

def build_session() -> requests.Session:
    """
    Establish a ColdFusion session with WebFIRE (3-step flow) and return
    a ready-to-use requests.Session.
    """
    s = requests.Session()
    s.headers["User-Agent"] = _UA
    s.headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"

    s.get(BASE_URL + "esearch.cfm", timeout=30)
    s.headers["Referer"] = BASE_URL + "esearch.cfm"
    s.post(BASE_URL + "esearch2.cfm",
           data={"reporttype": "All", "Submit": "Submit Search"},
           timeout=60)
    s.headers["Referer"] = BASE_URL + "esearch2.cfm"
    return s


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

def search(session: requests.Session, params: dict) -> list:
    """
    POST the search form and return a list of report dicts.
    params keys: organization, facility, startdate, enddate, state,
                 county, city, zip, CFRpart, CFRSubpart
    """
    payload = {
        "organization": params.get("organization", ""),
        "facility":     params.get("facility",     ""),
        "startdate":    params.get("startdate",    ""),
        "enddate":      params.get("enddate",      ""),
        "state":        params.get("state",        "AA"),
        "county":       params.get("county",       ""),
        "city":         params.get("city",         ""),
        "zip":          params.get("zip",          ""),
        "CFRpart":      params.get("CFRpart",      "All"),
        "CFRSubpart":   params.get("CFRSubpart",   ""),
        "FRS":          "",
        "Submit":       "Submit Search",
    }
    r = session.post(BASE_URL + "eSearchResults.cfm", data=payload, timeout=60)
    r.raise_for_status()
    return _parse_results(r.text)


def _parse_results(html: str) -> list:
    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table", class_="cell-border")
    if not table:
        return []

    seen = {}
    for row in table.find_all("tr")[2:]:
        cells = row.find_all("td")
        if len(cells) < 11:
            continue
        link = cells[10].find("a", href=True)
        if not link or "ID=" not in link["href"]:
            continue

        doc_id = link["href"].split("ID=")[-1].strip()
        if doc_id not in seen:
            seen[doc_id] = {
                "id":             doc_id,
                "organization":   cells[0].get_text(strip=True),
                "facility":       cells[1].get_text(strip=True),
                "city":           cells[2].get_text(strip=True),
                "state":          cells[3].get_text(strip=True),
                "county":         cells[4].get_text(strip=True),
                "date":           cells[5].get_text(strip=True),
                "report_type":    cells[6].get_text(strip=True),
                "report_subtype": cells[7].get_text(strip=True),
                "pollutants":     cells[8].get_text(strip=True),
                "filename":       link.get("title", ""),
                "downloaded":     False,
                "scanned":        False,
            }
    return list(seen.values())


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_report(session: requests.Session, doc_id: str,
                    download_dir: Path) -> tuple:
    """
    Download a single report ZIP.
    Returns (success: bool, path: Path|None, error: str).
    """
    dest = download_dir / f"{doc_id}.zip"
    if dest.exists():
        return True, dest, "already cached"

    try:
        r = session.get(DETAIL_URL, params={"ID": doc_id}, timeout=120)
        r.raise_for_status()
        # ZIP files start with PK (0x50 0x4B). Detect bad responses early.
        if not r.content[:2] == b"PK":
            content_type = r.headers.get("Content-Type", "unknown")
            header = r.content[:8].hex()
            return False, None, (
                f"WebFIRE returned non-ZIP content "
                f"(Content-Type: {content_type}, header bytes: {header})"
            )
        dest.write_bytes(r.content)
        return True, dest, ""
    except Exception as exc:
        return False, None, str(exc)


# ---------------------------------------------------------------------------
# Facility metadata extraction
# ---------------------------------------------------------------------------

def extract_file_meta(zip_path: Path) -> dict:
    """
    Extract facility name, city, and state directly from the report file.
    Used when no WebFIRE search result is available (e.g. Scan Local Folder).

    Returns a dict with keys 'facility', 'city', 'state' (all strings, may be empty).
    """
    empty = {"facility": "", "city": "", "state": ""}
    try:
        with zipfile.ZipFile(zip_path) as z:
            names = z.namelist()

            # ── ST: read qryEFFacility from XML ──────────────────────────────
            xml_names = [n for n in names if n.endswith(".xml")]
            if xml_names:
                try:
                    root = ET.fromstring(z.read(xml_names[0]))
                    fac = root.find("qryEFFacility")
                    if fac is not None:
                        return {
                            "facility": fac.findtext("Facility", "").strip(),
                            "city":     fac.findtext("City",     "").strip(),
                            "state":    fac.findtext("State",    "").strip(),
                        }
                except ET.ParseError:
                    pass
                return empty

            # ── AER: read facility sheet from Excel ──────────────────────────
            xlsx_names = [n for n in names if n.endswith((".xlsx", ".xlsm"))]
            if xlsx_names:
                import openpyxl, io as _io
                wb = openpyxl.load_workbook(
                    _io.BytesIO(z.read(xlsx_names[0])),
                    read_only=True, data_only=True)
                try:
                    # Templates use either "Company_Information" or
                    # "Facility Information" depending on the subpart.
                    fac_sheet = next(
                        (s for s in wb.sheetnames
                         if s in ("Company_Information", "Facility Information")),
                        None)
                    if fac_sheet is None:
                        return empty
                    ws = wb[fac_sheet]
                    # Scan for the anchor row that contains field-name tokens
                    # (e.g. "CompanyName", "FacilityName") then read the first
                    # real data row (skipping "e.g.:" example rows).
                    NAME_KEYS  = {"companyname", "facilityname"}
                    CITY_KEYS  = {"cityname"}
                    STATE_KEYS = {"statename"}
                    col_company = col_city = col_state = None
                    in_data = False
                    for row in ws.iter_rows(values_only=True):
                        row = list(row)
                        if not in_data:
                            # Detect anchor row: any cell whose lowercase value
                            # is in our known key sets
                            hits = {
                                str(c).lower().strip()
                                for c in row if isinstance(c, str)
                            }
                            if hits & (NAME_KEYS | CITY_KEYS | STATE_KEYS):
                                for i, c in enumerate(row):
                                    if not isinstance(c, str):
                                        continue
                                    cl = c.lower().strip()
                                    if cl in NAME_KEYS:
                                        col_company = i
                                    elif cl in CITY_KEYS:
                                        col_city = i
                                    elif cl in STATE_KEYS:
                                        col_state = i
                                in_data = True
                            continue
                        # Skip example rows and blank rows
                        if not any(c is not None for c in row):
                            continue
                        if any(isinstance(c, str) and
                               c.strip().lower().startswith("e.g.") for c in row):
                            continue
                        # First real data row
                        def _get(idx):
                            if idx is None or idx >= len(row):
                                return ""
                            v = row[idx]
                            return str(v).strip() if v is not None else ""
                        name = _get(col_company)
                        if name:
                            return {
                                "facility": name,
                                "city":     _get(col_city),
                                "state":    _get(col_state),
                            }
                        break
                finally:
                    wb.close()
    except Exception:
        pass
    return empty


# ---------------------------------------------------------------------------
# Report classifier
# ---------------------------------------------------------------------------

def classify_report(zip_path: Path) -> str:
    """
    Inspect a downloaded file and return one of:
      'ST'           — XML-based stack test report
      'AER'          — Excel-based CEDRI compliance report
      'PDF'          — Valid ZIP containing only PDF files (no structured data)
      'BADDOWNLOAD_PDF' — File is not a ZIP; WebFIRE returned a bare PDF instead
      'BADDOWNLOAD'  — File is not a ZIP and content type is unrecognised
      'UNKNOWN'      — Cannot determine type
    """
    try:
        with zipfile.ZipFile(zip_path) as z:
            names = z.namelist()
            exts = {Path(n).suffix.lower() for n in names if Path(n).suffix}
            if ".xml" in exts:
                return "ST"
            if ".xlsx" in exts or ".xlsm" in exts:
                return "AER"
            if exts and all(e == ".pdf" for e in exts):
                return "PDF"
            return "UNKNOWN"
    except zipfile.BadZipFile:
        # The file is not a ZIP — check what WebFIRE actually returned.
        try:
            with open(zip_path, "rb") as f:
                if f.read(4) == b"%PDF":
                    return "BADDOWNLOAD_PDF"
        except OSError:
            pass
        return "BADDOWNLOAD"
    except Exception:
        return "UNKNOWN"


# ---------------------------------------------------------------------------
# AER scanning helpers
# ---------------------------------------------------------------------------

def _extract_citation(ws) -> str:
    """
    Extract CitationID from the standard CEDRI metadata block.
    Looks at row 3, column C (0-indexed: row 2, col 2).
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == _CITATION_ROW:
            val = row[_CITATION_COL] if len(row) > _CITATION_COL else None
            return str(val).strip() if val is not None else ""
        if i > 6:
            break
    return ""


def _get_target_sheets(citation: str, sheet_names: list) -> list:
    """
    Return an ordered list of sheet names to scan given a CitationID.
    Filters to sheets that actually exist in the workbook.
    """
    citation_lower = citation.lower()
    for key, sheets in CITATION_TO_SHEETS.items():
        if key.lower() in citation_lower:
            return [s for s in sheets if s in sheet_names]
    return []


def _is_summary_field(name: str) -> bool:
    """
    Return True if this field name is a numeric deviation summary/count column.
    A non-zero value in such a field indicates a deviation.
    """
    n = name.lower()
    # Direct substring matches from the keyword list
    if any(kw in n for kw in _SUMMARY_FIELD_KEYWORDS):
        return True
    # Field ends in _Number and contains "deviation" or "count" → count-based
    if n.endswith("_number") and ("deviat" in n or "count" in n or "malfunction" in n):
        return True
    return False


def _coerce_float(val) -> float | None:
    """Coerce a cell value to float; return None on failure."""
    if val is None:
        return None
    try:
        return float(str(val).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


# Field names that are accepted as the anchor (column B) in the machine-readable row.
# Most CEDRI templates use RecordId; JJJJJJ uses EngineId; others may differ.
_ANCHOR_FIELD_NAMES = frozenset({
    "RecordId", "EngineId", "CompanyRecordId", "SiteId", "FacilityId",
})


def _find_data_start(rows: list) -> tuple:
    """
    Locate the CEDRI header anchor row.

    Looks for a row where column B (index 1) contains a known machine-readable
    identifier field name (typically 'RecordId', sometimes 'EngineId' etc.).
    Falls back to a structural heuristic for templates that use other identifiers.

    Returns:
        (anchor_idx, field_names_dict, data_start_idx)
        anchor_idx: row index of the machine-readable field name row
        field_names_dict: {col_index: field_name_str}
        data_start_idx: first row index of actual data (after e.g. examples)
    Returns (None, {}, None) if no anchor is found.
    """
    anchor_idx = None

    # Primary: exact match on known anchor field names in column B
    for i, row in enumerate(rows):
        if len(row) > 1 and str(row[1] or "").strip() in _ANCHOR_FIELD_NAMES:
            anchor_idx = i
            break

    # Fallback: scan rows 10-16 for a row where column B looks like a
    # machine-readable field name (camelCase/PascalCase, no spaces)
    if anchor_idx is None:
        for i, row in enumerate(rows[8:18], start=8):
            if len(row) < 2:
                continue
            val = str(row[1] or "").strip()
            if (val and " " not in val and "\n" not in val
                    and len(val) >= 2 and len(val) <= 40
                    and (any(c.isupper() for c in val) or "_" in val)
                    and not str(row[1] or "").strip().lower().startswith("e.g.")):
                anchor_idx = i
                break

    if anchor_idx is None:
        return None, {}, None

    field_row = rows[anchor_idx]
    field_names = {}
    for j, v in enumerate(field_row):
        if v is not None and str(v).strip():
            field_names[j] = str(v).strip()

    # Advance past example rows (cells starting with "e.g.")
    data_start = anchor_idx + 1
    for i in range(anchor_idx + 1, len(rows)):
        row = rows[i]
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            continue
        if all(str(v).strip().lower().startswith("e.g.") for v in non_null):
            data_start = i + 1
        else:
            data_start = i
            break

    return anchor_idx, field_names, data_start


def _format_summary_description(n_rows: int, summary_hits: dict) -> str:
    """
    Convert summary_hits {field_name: max_value} into a human-readable string.

    Lookup strategy:
      1. Exact match on field_name.lower()
      2. Strip trailing sheet-identifier suffix (_EES, _EED, _CMS, _CMSDo, etc.)
         and try again — handles CEDRI templates that append sheet acronyms to field names
      3. Fall back to 'value (FieldName)'
    """
    parts = []
    for field_name, value in list(summary_hits.items())[:5]:
        n = field_name.lower()
        template = _FIELD_FRIENDLY.get(n)
        # Strip one trailing _XYZ suffix and retry
        if template is None and "_" in n:
            stripped = n.rsplit("_", 1)[0]
            template = _FIELD_FRIENDLY.get(stripped)
        if template:
            parts.append(template.format(v=value))
        else:
            parts.append(f"{value:g} ({field_name})")
    desc = "; ".join(parts)
    if n_rows > 1:
        desc = f"{n_rows} rows — " + desc
    return desc


def _combine_datetime(date_val, time_val):
    """
    Combine a date cell and a separate time cell (as returned by openpyxl) into
    a single Python datetime.  Returns None if the date cannot be parsed.
    """
    import datetime as _dt
    if date_val is None:
        return None

    # --- date part ---
    if isinstance(date_val, _dt.datetime):
        base = date_val.replace(hour=0, minute=0, second=0, microsecond=0)
    elif isinstance(date_val, _dt.date):
        base = _dt.datetime(date_val.year, date_val.month, date_val.day)
    else:
        try:
            base = _dt.datetime.strptime(str(date_val).strip()[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            return None

    # --- time part ---
    if time_val is None:
        return base
    if isinstance(time_val, _dt.time):
        return base + _dt.timedelta(
            hours=time_val.hour, minutes=time_val.minute, seconds=time_val.second)
    if isinstance(time_val, _dt.timedelta):
        return base + time_val
    if isinstance(time_val, (int, float)):
        # openpyxl may return fraction-of-day floats for time-only cells
        total_sec = int(time_val * 86400)
        return base + _dt.timedelta(seconds=total_sec)
    if isinstance(time_val, str):
        try:
            parts = time_val.strip().split(":")
            h, m = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
            s = int(float(parts[2])) if len(parts) > 2 else 0
            return base + _dt.timedelta(hours=h, minutes=m, seconds=s)
        except (ValueError, IndexError):
            pass
    return base


def _scan_fenceline_sheet(ws_rows: list, report_meta: dict, citation: str) -> list:
    """
    Analyze a §63.655(h)(8) fenceline monitoring report.

    Algorithm (per §63.655(h)(3)):
      - For each Regular Monitor sampler, compute the annual average benzene
        concentration across all sampling periods.
      - Flag YES if any location's average exceeds FENCELINE_THRESHOLD_UGM3.
    Returns a list containing one _aer_row dict.
    """
    anchor_idx, field_names, data_start = _find_data_start(ws_rows)
    if anchor_idx is None:
        return [_aer_error_row(report_meta, citation,
                               "Sample Results: cannot locate header anchor")]

    # Build column-name → index map (case-insensitive)
    col_of = {v.lower(): k for k, v in field_names.items()}
    sampler_name_col  = col_of.get("samplername")
    sampler_type_col  = col_of.get("samplertype")
    benzene_col       = col_of.get("benzeneamt")

    if any(c is None for c in (sampler_name_col, sampler_type_col, benzene_col)):
        return [_aer_error_row(report_meta, citation,
                               "Sample Results: missing expected columns "
                               "(SamplerName, SamplerType, BenzeneAmt)")]

    # Collect readings for Regular Monitor samplers only
    readings: dict[str, list[float]] = {}
    for row in ws_rows[data_start:]:
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            continue
        if all(str(v).strip().lower().startswith("e.g.") for v in non_null):
            continue

        def _cell(col):
            return row[col] if col < len(row) else None

        stype = str(_cell(sampler_type_col) or "").strip()
        if "regular monitor" not in stype.lower():
            continue

        sname   = str(_cell(sampler_name_col) or "").strip()
        benzene = _coerce_float(_cell(benzene_col))
        if sname and benzene is not None:
            readings.setdefault(sname, []).append(benzene)

    if not readings:
        return [_aer_row(report_meta, citation, "Sample Results", "no",
                         "No Regular Monitor readings found in Sample Results")]

    # Compute annual average per sampler and compare to threshold
    exceedances = []
    sampler_avgs = []
    for sname, vals in sorted(readings.items()):
        avg = sum(vals) / len(vals)
        sampler_avgs.append((sname, avg, len(vals)))
        if avg > FENCELINE_THRESHOLD_UGM3:
            exceedances.append(f"{sname}: {avg:.3f} µg/m³")

    n_locs  = len(readings)
    max_avg = max(avg for _, avg, _ in sampler_avgs)

    if exceedances:
        desc = (f"ACTION LEVEL EXCEEDED at {len(exceedances)} of {n_locs} location(s). "
                f"{'; '.join(exceedances[:4])}. "
                f"Threshold: {FENCELINE_THRESHOLD_UGM3} µg/m³ annual average.")
        return [_aer_row(report_meta, citation, "Sample Results", "YES", desc)]
    else:
        desc = (f"No exceedance — {n_locs} location(s) monitored; "
                f"highest annual avg: {max_avg:.3f} µg/m³ "
                f"(threshold: {FENCELINE_THRESHOLD_UGM3} µg/m³)")
        return [_aer_row(report_meta, citation, "Sample Results", "no", desc)]


def _scan_turbine_sheet(ws_rows: list, report_meta: dict, citation: str) -> list:
    """
    Analyze a §60.4214(d)(3) stationary combustion turbine annual report.

    Algorithm:
      - For each engine, sum the duration of all non-emergency use events
        (EndDate+EndTime − StartDate+StartTime) in hours.
      - Flag YES if any engine exceeds TURBINE_ANNUAL_LIMIT_HRS.
    Returns one _aer_row per engine.
    """
    anchor_idx, field_names, data_start = _find_data_start(ws_rows)
    if anchor_idx is None:
        return [_aer_error_row(report_meta, citation,
                               "Non-emergency Use: cannot locate header anchor")]

    col_of = {v.lower(): k for k, v in field_names.items()}
    eng_col  = col_of.get("engineid")
    sd_col   = col_of.get("nonemergencystartdate")
    st_col   = col_of.get("nonemergencystarttime")
    ed_col   = col_of.get("nonemergencyenddate")
    et_col   = col_of.get("nonemergencyendtime")

    if any(c is None for c in (eng_col, sd_col, st_col, ed_col, et_col)):
        return [_aer_error_row(report_meta, citation,
                               "Non-emergency Use: missing expected columns")]

    engine_hours: dict[str, float] = {}

    for row in ws_rows[data_start:]:
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            continue
        if all(str(v).strip().lower().startswith("e.g.") for v in non_null):
            continue

        def _cell(col):
            return row[col] if col < len(row) else None

        eng     = str(_cell(eng_col) or "").strip()
        start   = _combine_datetime(_cell(sd_col), _cell(st_col))
        end     = _combine_datetime(_cell(ed_col), _cell(et_col))

        if not eng or start is None or end is None:
            continue
        dur_hrs = max(0.0, (end - start).total_seconds() / 3600.0)
        engine_hours[eng] = engine_hours.get(eng, 0.0) + dur_hrs

    if not engine_hours:
        return [_aer_row(report_meta, citation, "Non-emergency Use", "no",
                         "No non-emergency use events recorded")]

    results = []
    for eng, hrs in sorted(engine_hours.items()):
        if hrs > TURBINE_ANNUAL_LIMIT_HRS:
            desc = (f"Engine {eng}: {hrs:.1f} hrs of non-emergency use — "
                    f"EXCEEDS {TURBINE_ANNUAL_LIMIT_HRS:.0f}-hr annual limit")
            results.append(_aer_row(report_meta, citation, "Non-emergency Use", "YES", desc))
        else:
            desc = (f"Engine {eng}: {hrs:.1f} hrs of non-emergency use — "
                    f"within {TURBINE_ANNUAL_LIMIT_HRS:.0f}-hr annual limit")
            results.append(_aer_row(report_meta, citation, "Non-emergency Use", "no", desc))
    return results


def _scan_aer_sheet(rows: list, sheet_name: str) -> dict:
    """
    Scan one worksheet (provided as a list of row tuples) for deviation data.

    Returns a dict:
        deviation:   "YES" | "no" | "manual-review"
        description: human-readable summary
    """
    # Special case: Voltage Frequency Deviation (JJJJJJ) — non-standard layout
    # Headers at row 4 (0-indexed 3), data at row 5 (0-indexed 4)
    if sheet_name == "Voltage Frequency Deviation":
        data_rows = []
        for row in rows[4:]:
            non_null = [v for v in row if v is not None and str(v).strip()]
            if non_null:
                data_rows.append(row)
        if data_rows:
            return {
                "deviation": "YES",
                "description": f"{len(data_rows)} voltage/frequency deviation event(s)",
            }
        return {"deviation": "no", "description": ""}

    # Standard CEDRI anchor detection
    anchor_idx, field_names, data_start = _find_data_start(rows)

    if anchor_idx is None:
        # No RecordId anchor — try a shallow heuristic scan of all rows
        return _scan_aer_sheet_heuristic(rows, sheet_name)

    # Collect actual data rows (non-null, non-example)
    data_rows = []
    for row in rows[data_start:]:
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            continue
        if all(str(v).strip().lower().startswith("e.g.") for v in non_null):
            continue
        data_rows.append(row)

    if not data_rows:
        return {"deviation": "no", "description": ""}

    # Check for non-zero values in numeric deviation-summary fields
    summary_hits = {}
    for row in data_rows:
        for col_idx, field_name in field_names.items():
            if col_idx >= len(row):
                continue
            if not _is_summary_field(field_name):
                continue
            num = _coerce_float(row[col_idx])
            if num is not None and num > 0:
                if field_name not in summary_hits:
                    summary_hits[field_name] = 0.0
                summary_hits[field_name] = max(summary_hits[field_name], num)

    if summary_hits:
        return {
            "deviation": "YES",
            "description": _format_summary_description(len(data_rows), summary_hits),
        }

    # Determine whether this is a summary or event-level sheet
    has_summary_fields = any(_is_summary_field(n) for n in field_names.values())
    if has_summary_fields:
        # Summary sheet with all-zero values → no deviation
        return {
            "deviation": "no",
            "description": f"{len(data_rows)} row(s), all deviation totals are zero",
        }

    # Event-level sheet: any data row = a deviation event
    desc_parts = _first_row_description(data_rows[0], field_names)
    desc = f"{len(data_rows)} deviation event(s)"
    if desc_parts:
        desc += " — " + "; ".join(desc_parts)
    return {"deviation": "YES", "description": desc}


def _scan_aer_sheet_heuristic(rows: list, sheet_name: str) -> dict:
    """
    Fallback scanner for sheets without a RecordId anchor.
    Looks for any non-empty rows past a short header block (first 20 rows).
    """
    data_rows = []
    for row in rows[20:]:
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            continue
        if all(str(v).strip().lower().startswith("e.g.") for v in non_null):
            continue
        data_rows.append(row)

    if data_rows:
        return {
            "deviation": "YES",
            "description": (
                f"{len(data_rows)} data row(s) found "
                "(no RecordId anchor — heuristic detection)"
            ),
        }
    return {"deviation": "no", "description": "No data rows found (heuristic)"}


def _first_row_description(row: tuple, field_names: dict) -> list:
    """Extract up to 3 key field values from the first data row for a description."""
    key_fields = {
        "deviationdesc", "description", "deviationtype", "excessreason",
        "deviationcause", "processunit", "processunit_dnc",
        "deviationfactype", "deviationfacid", "pollutant",
    }
    parts = []
    for col_idx, fname in field_names.items():
        if fname.lower() not in key_fields:
            continue
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None:
            continue
        v_str = str(v).strip()
        if not v_str or v_str.lower().startswith("e.g."):
            continue
        parts.append(f"{fname}={v_str[:60]}")
        if len(parts) >= 3:
            break
    return parts


# ---------------------------------------------------------------------------
# AER report scanner
# ---------------------------------------------------------------------------

def scan_aer_report(local_path: Path, report_meta: dict) -> list:
    """
    Scan one AER (Excel-based CEDRI compliance report) for deviations.
    Returns a list of finding dicts — one per deviation-relevant sheet scanned.
    """
    try:
        import openpyxl
    except ImportError:
        return [_aer_error_row(
            report_meta, "",
            "openpyxl not installed — run: pip install openpyxl"
        )]

    # ── Open ZIP and load Excel bytes ───────────────────────────────────────
    try:
        with zipfile.ZipFile(local_path) as z:
            xl_names = [
                n for n in z.namelist()
                if Path(n).suffix.lower() in (".xlsx", ".xlsm")
            ]
            if not xl_names:
                return [_aer_error_row(report_meta, "", "No Excel file found in ZIP")]

            # Prefer the shallowest Excel (avoid nested ZIPs if any)
            xl_name = sorted(xl_names, key=lambda n: n.count("/"))[0]
            xl_bytes = io.BytesIO(z.read(xl_name))

    except zipfile.BadZipFile:
        return [_aer_error_row(report_meta, "", "Not a valid ZIP file")]
    except Exception as exc:
        return [_aer_error_row(report_meta, "", f"ZIP open error: {exc}")]

    # ── Load workbook ───────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(xl_bytes, read_only=True, data_only=True)
    except Exception as exc:
        return [_aer_error_row(report_meta, "", f"Excel open error: {exc}")]

    rows = []
    citation = ""

    try:
        sheet_names = wb.sheetnames

        # ── Phase 1: Extract CitationID ─────────────────────────────────────
        # Priority: Welcome sheet, then first sheet
        for ws_name in (["Welcome"] if "Welcome" in sheet_names else []) + [sheet_names[0]]:
            ws = wb[ws_name]
            citation = _extract_citation(ws)
            if citation:
                break

        # Special case: DDDDD Tune-Up report (no citation, only Instructions/Data)
        if not citation and set(sheet_names) <= {"Instructions", "Data", "Revisions"}:
            wb.close()
            return [_aer_row(
                report_meta, "", "N/A", "manual-review",
                "Tune-Up Report (§63.7550(c)(1)) — compliance data is in the attached PDF"
            )]

        # ── Phase 2: Route to target deviation sheets ───────────────────────
        citation_lower = citation.lower()

        # ── Fenceline monitoring (§63.655) — auto-detect via benzene averaging ──
        if "63.655" in citation_lower:
            sample_sheet = "Sample Results"
            if sample_sheet in sheet_names:
                ws = wb[sample_sheet]
                ws_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                return _scan_fenceline_sheet(ws_rows, report_meta, citation)
            wb.close()
            return [_aer_error_row(report_meta, citation,
                                   "Fenceline: 'Sample Results' sheet not found in workbook")]

        # ── Stationary turbine (§60.4214) — auto-detect via non-emergency hours ──
        if "60.4214" in citation_lower or "63.4214" in citation_lower:
            turbine_sheet = "Non-emergency Use"
            if turbine_sheet in sheet_names:
                ws = wb[turbine_sheet]
                ws_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                return _scan_turbine_sheet(ws_rows, report_meta, citation)
            wb.close()
            return [_aer_error_row(report_meta, citation,
                                   "Turbine: 'Non-emergency Use' sheet not found in workbook")]

        # ── NSPS OOOOb and OOOOa — still manual-review / no-table ──────────────
        if "60.5422b" in citation_lower:
            wb.close()
            return [_aer_row(
                report_meta, citation, "N/A", "manual-review",
                "NSPS OOOOb — leak records require domain-specific interpretation "
                "(not yet auto-scannable)"
            )]
        if "60.5422a" in citation_lower:
            wb.close()
            return [_aer_row(
                report_meta, citation, "N/A", "no",
                "NSPS OOOOa Semiannual — flat inventory report, "
                "no structured deviation table to scan"
            )]

        target_sheets = _get_target_sheets(citation, sheet_names)
        fallback_used = False
        unmatched_citation = ""

        # Keyword fallback when no routing entry matches the citation
        if not target_sheets:
            kw_sheets = [
                s for s in sheet_names
                if any(kw in s.lower() for kw in KEYWORD_FALLBACK)
            ]
            if kw_sheets:
                target_sheets = kw_sheets
                fallback_used = True
                unmatched_citation = citation

        if not target_sheets:
            wb.close()
            return [_aer_row(
                report_meta, citation, "—", "no",
                f"No deviation sheets identified "
                f"(citation: '{citation or 'not found'}'; "
                f"sheets: {', '.join(sheet_names[:8])})",
                fallback_used=False,
                unmatched_citation=citation if citation else "",
            )]

        # ── Phase 3: Scan each target sheet ────────────────────────────────
        for sheet_name in target_sheets:
            if sheet_name not in sheet_names:
                continue
            try:
                ws = wb[sheet_name]
                ws_rows = list(ws.iter_rows(values_only=True))
                finding = _scan_aer_sheet(ws_rows, sheet_name)
                rows.append(_aer_row(
                    report_meta, citation, sheet_name,
                    finding["deviation"], finding["description"],
                    fallback_used=fallback_used,
                    unmatched_citation=unmatched_citation,
                ))
            except Exception as exc:
                rows.append(_aer_error_row(
                    report_meta, citation,
                    f"Sheet '{sheet_name}': {exc}"
                ))

        if not rows:
            rows.append(_aer_row(
                report_meta, citation, "—", "no",
                "Target sheets found but no data extracted"
            ))

    except Exception as exc:
        rows.append(_aer_error_row(report_meta, citation, f"Scan error: {exc}"))
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return rows


def _aer_row(meta: dict, citation: str, sheet: str,
             deviation: str, description: str,
             fallback_used: bool = False,
             unmatched_citation: str = "") -> dict:
    """
    Build an AER finding row.

    Reuses the existing ST display schema for GUI compatibility:
      location  → sheet name
      pollutant → citation ID  (column repurposed for AER display)
      regulation → full citation string
      error      → description / notes (shown in Notes column)
    AER-specific keys (citation, sheet, description) are extras for CSV export.
    fallback_used / unmatched_citation track when keyword routing was used
    because the CitationID was not found in the routing table.
    """
    return {
        # Shared keys (used by both ST and AER in GUI + CSV)
        "report_id":          meta["id"],
        "facility":           meta["facility"],
        "city":               meta["city"],
        "state":              meta["state"],
        "date":               meta["date"],
        "report_type":        "AER",
        "location":           sheet,
        "pollutant":          citation,
        "unit":               "",
        "n_runs":             "",
        "avg_measured":       "",
        "limit":              "",
        "pct_of_limit":       "",
        "regulation":         citation,
        "deviation":          deviation,
        "error":              description,
        # AER-specific extras (written to CSV, ignored by ST display path)
        "citation":           citation,
        "sheet":              sheet,
        "description":        description,
        # Citation diagnostics (item 3)
        "fallback_used":      fallback_used,
        "unmatched_citation": unmatched_citation if fallback_used else "",
    }


def _aer_error_row(meta: dict, citation: str, msg: str) -> dict:
    """Build an AER error row."""
    return _aer_row(meta, citation, "", "error", msg)


# ---------------------------------------------------------------------------
# Stack Test (ST) XML scanner — unchanged logic, refactored into helper
# ---------------------------------------------------------------------------

def _elem_text(elem, tag: str) -> str:
    child = elem.find(tag)
    return (child.text or "").strip() if child is not None else ""


def _to_float(s: str):
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _unit_key(unit_str: str) -> str:
    """Normalize a unit string for fuzzy matching across XML formats."""
    u = unit_str.lower().strip()
    # Common variations seen across WebFIRE XML formats
    u = u.replace("grains", "gr").replace("grain", "gr")
    u = u.replace(" corrected", "").replace("@15%o2", "").replace("@15% o2", "")
    u = u.replace("(", "").replace(")", "").replace("%", "pct").replace(" ", "")
    return u


def _scan_st_tbl_emis(root, tbl_concs: list, report_meta: dict) -> list:
    """
    Parse stack test XML in the tblEmisConcs format (newer CEDRI schema).

    This format stores one measurement per element in tblEmisConcs:
      Location   — stack/source location
      Method     — EPA test method used
      EmisConc   — unit string (e.g. 'percent(%)', 'grains/dscf', 'ppm corrected')
      CorrPerc   — corrected/normalized measured value (numeric string)

    Regulatory limits are in qryEFRegs:
      PollName       — pollutant name
      Limit          — numeric limit
      EmissionUnit   — unit string
      Part_SubPart   — regulation reference

    Matching is done by normalizing both unit strings and finding the closest match.
    """
    # Build limit lookup: normalized_unit → (poll, lim, reg, raw_unit)
    reg_by_unit = {}
    for elem in root.findall("qryEFRegs"):
        poll = _elem_text(elem, "PollName")
        lim  = _to_float(_elem_text(elem, "Limit"))
        unit = _elem_text(elem, "EmissionUnit")
        reg  = _elem_text(elem, "Part_SubPart")
        if lim is not None and lim > 0 and unit:
            key = _unit_key(unit)
            if key not in reg_by_unit:
                reg_by_unit[key] = (poll, lim, reg, unit)

    rows = []
    seen = set()
    for elem in tbl_concs:
        loc    = _elem_text(elem, "Location")
        method = _elem_text(elem, "Method")
        unit   = _elem_text(elem, "EmisConc")
        val    = _to_float(_elem_text(elem, "CorrPerc"))

        if val is None or not unit:
            continue

        # Match to a limit by unit key
        key = _unit_key(unit)
        match = reg_by_unit.get(key)

        # Fallback: check if any limit unit is a prefix of the measurement unit
        if match is None:
            for lkey, lval in reg_by_unit.items():
                if key.startswith(lkey[:4]) or lkey.startswith(key[:4]):
                    match = lval
                    break

        poll  = match[0] if match else method
        lim   = match[1] if match else None
        reg   = match[2] if match else ""
        lunit = match[3] if match else unit

        dedup_key = (loc, poll, lunit)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        rows.append({
            "report_id":    report_meta["id"],
            "facility":     report_meta["facility"],
            "city":         report_meta["city"],
            "state":        report_meta["state"],
            "date":         report_meta["date"],
            "report_type":  report_meta.get("report_type", "ST"),
            "location":     loc,
            "pollutant":    poll,
            "unit":         lunit,
            "n_runs":       1,
            "avg_measured": round(val, 6),
            "limit":        lim if lim is not None else "",
            "pct_of_limit": round(val / lim * 100, 1) if lim else "",
            "regulation":   reg,
            "deviation":    ("YES" if val > lim else "no") if lim is not None else "no limit",
            "error":        "",
        })

    if not rows:
        return [_error_row(report_meta,
                           "tblEmisConcs format — no parseable emission measurements found")]
    return rows


def _scan_st_report(local_path: Path, report_meta: dict) -> list:
    """Scan a Stack Test (XML) report for emission vs. limit comparisons."""
    try:
        outer = zipfile.ZipFile(local_path)
    except zipfile.BadZipFile:
        return [_error_row(report_meta, "Not a valid ZIP file")]

    xml_names = [n for n in outer.namelist() if n.endswith(".xml")]
    if not xml_names:
        exts = {Path(n).suffix.lower() for n in outer.namelist()}
        return [_error_row(
            report_meta,
            f"No XML found (contains: {', '.join(sorted(exts)) or 'nothing'})"
        )]

    try:
        root = ET.fromstring(outer.read(xml_names[0]))
    except ET.ParseError as exc:
        return [_error_row(report_meta, f"XML parse error: {exc}")]

    # ── MATS Transition wrapper: XML is metadata only; data is in PDF ────────
    if root.tag == "MatsTransitionMetaData":
        row = _error_row(report_meta, "")
        row["deviation"] = "manual-review"
        row["error"] = (
            "MATS Transition report — stack test data is in the attached PDF "
            "(XML contains only submission metadata)"
        )
        return [row]

    # ── Regulatory limits keyed by (pds_id, poll_name, unit) ────────────────
    limits = {}
    global_limits = {}
    for elem in root.findall("qryEFRegs"):
        pds  = _elem_text(elem, "PDSid")
        poll = _elem_text(elem, "PollName")
        unit = _elem_text(elem, "EmissionUnit")
        lim  = _to_float(_elem_text(elem, "Limit"))
        reg  = _elem_text(elem, "Part_SubPart")
        # Limit of 0 means "not populated" in this XML format — skip it so the
        # measurement is reported as "no limit" rather than a false deviation.
        if lim is not None and lim > 0:
            limits[(pds, poll, unit)] = lim
            if (poll, unit) not in global_limits:
                global_limits[(poll, unit)] = (reg, lim)

    # ── Per-run emission measurements (legacy qryEFEmis format) ─────────────
    runs = defaultdict(list)
    for elem in root.findall("qryEFEmis"):
        pds  = _elem_text(elem, "PDSid")
        loc  = _elem_text(elem, "Location")
        poll = _elem_text(elem, "PollName")
        unit = _elem_text(elem, "EmissionUnit")
        val  = _to_float(_elem_text(elem, "EmissionValue"))
        if val is not None:
            runs[(pds, loc, poll, unit)].append(val)

    # ── tblEmisConcs format (newer format: opacity, concentration tests) ─────
    # Used when qryEFEmis is absent; matches measurements to limits by unit.
    tbl_concs = root.findall("tblEmisConcs")
    if not runs and tbl_concs:
        return _scan_st_tbl_emis(root, tbl_concs, report_meta)

    if not runs:
        # dataroot format with process-params only (no emission concentrations)
        if root.tag == "dataroot" and root.find("tblProcessRunData") is not None:
            row = _error_row(report_meta, "")
            row["deviation"] = "manual-review"
            row["error"] = (
                "Process parameters only — XML records run conditions but no "
                "pollutant concentration measurements vs. limits"
            )
            return [row]
        return [_error_row(report_meta, "No emission data in XML")]

    rows = []
    for (pds, loc, poll, unit), vals in runs.items():
        avg = sum(vals) / len(vals)

        lim = limits.get((pds, poll, unit))
        reg = ""
        if lim is None:
            match = global_limits.get((poll, unit))
            if match:
                reg, lim = match
        else:
            for elem in root.findall("qryEFRegs"):
                if (_elem_text(elem, "PDSid") == pds and
                        _elem_text(elem, "PollName") == poll and
                        _elem_text(elem, "EmissionUnit") == unit):
                    reg = _elem_text(elem, "Part_SubPart")
                    break

        rows.append({
            "report_id":    report_meta["id"],
            "facility":     report_meta["facility"],
            "city":         report_meta["city"],
            "state":        report_meta["state"],
            "date":         report_meta["date"],
            "report_type":  report_meta["report_type"],
            "location":     loc,
            "pollutant":    poll,
            "unit":         unit,
            "n_runs":       len(vals),
            "avg_measured": round(avg, 6),
            "limit":        lim if lim is not None else "",
            "pct_of_limit": round(avg / lim * 100, 1) if lim else "",
            "regulation":   reg,
            "deviation":    ("YES" if avg > lim else "no") if lim is not None else "no limit",
            "error":        "",
        })

    return rows


# ---------------------------------------------------------------------------
# Public scan entry point
# ---------------------------------------------------------------------------

def scan_report(local_path: Path, report_meta: dict) -> list:
    """
    Classify and scan one downloaded report ZIP for compliance deviations.
    Dispatches to the XML scanner (ST) or Excel scanner (AER) automatically.
    Returns a list of finding dicts (one per comparison or sheet).
    """
    rtype = classify_report(local_path)

    if rtype == "AER":
        return scan_aer_report(local_path, report_meta)

    if rtype == "PDF":
        return [_error_row(
            report_meta,
            "PDF-only report — no structured deviation data (manual review required)"
        )]

    if rtype == "BADDOWNLOAD_PDF":
        row = _error_row(report_meta, "")
        row["deviation"] = "manual-review"
        row["error"] = (
            "WebFIRE returned a PDF instead of a ZIP — "
            "review the PDF document manually."
        )
        return [row]

    if rtype == "BADDOWNLOAD":
        row = _error_row(report_meta, "")
        row["deviation"] = "error"
        row["error"] = (
            "Bad download — WebFIRE returned a non-ZIP file. Re-download this report."
        )
        return [row]

    # ST or UNKNOWN: attempt XML scan (existing logic)
    return _scan_st_report(local_path, report_meta)


def _error_row(meta: dict, msg: str) -> dict:
    return {
        "report_id":    meta["id"],
        "facility":     meta["facility"],
        "city":         meta["city"],
        "state":        meta["state"],
        "date":         meta["date"],
        "report_type":  meta["report_type"],
        "location":     "",
        "pollutant":    "",
        "unit":         "",
        "n_runs":       "",
        "avg_measured": "",
        "limit":        "",
        "pct_of_limit": "",
        "regulation":   "",
        "deviation":    "error",
        "error":        msg,
    }
